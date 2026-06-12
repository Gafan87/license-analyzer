# modules/capacity_mapper.py
"""
Модуль для загрузки описаний CapacityKey из Excel файлов
Файлы: license_details/{domain}_license_details.xlsx
Поиск: сначала в сетевом хранилище, потом локально
Поддерживает маппинг NE Type -> Domain -> Файл описаний
"""

import os
import openpyxl
from modules.logger import get_logger
from flask import current_app
import math

logger = get_logger(__name__)

# Кэш описаний (загружается один раз)
_capacity_descriptions_cache = {}

# Кэш для маппинга NE Type -> Domain -> Файл
_ne_type_mapping_cache = None

def load_ne_type_mapping(mapping_file=None):
    """
    Загружает маппинг NE Type -> Domain -> Файл описаний из Excel
    """
    global _ne_type_mapping_cache
    
    if _ne_type_mapping_cache is not None:
        return _ne_type_mapping_cache
    
    # Если путь не передан, пытаемся получить из конфига
    if mapping_file is None:
        try:
            from flask import current_app
            mapping_file = current_app.config.get('ne_type_mapping_file')
        except (ImportError, RuntimeError):
            # Вне контекста Flask, пробуем прочитать config.json напрямую
            import json
            try:
                with open('config.json', 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    mapping_file = config.get('ne_type_mapping_file')
            except:
                mapping_file = None
    
    # Если путь всё ещё не определён, используем стандартный
    if not mapping_file:
        mapping_file = 'mapping/ne_type_mapping.xlsx'
    
    # Пробуем разные варианты пути
    if not os.path.exists(mapping_file):
        # Пробуем с сетевым путём
        alt_path1 = r'\\A00742028-C5NC5\_Licenses\mapping\ne_type_mapping.xlsx'
        if os.path.exists(alt_path1):
            mapping_file = alt_path1
    
    if not os.path.exists(mapping_file):
        # Пробуем с прямыми слешами
        alt_path2 = '//A00742028-C5NC5/_Licenses/mapping/ne_type_mapping.xlsx'
        if os.path.exists(alt_path2):
            mapping_file = alt_path2
    
    if not os.path.exists(mapping_file):
        logger.warning(f"Файл маппинга NE Type не найден: {mapping_file}")
        return {}
    
    try:
        wb = openpyxl.load_workbook(mapping_file, data_only=True)
        sheet = wb.active
        
        mapping = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            
            ne_type = str(row[0]).strip() if row[0] else ''
            domain = str(row[1]).strip() if row[1] else ''
            details_file = str(row[2]).strip() if row[2] else ''
            sheet_name = str(row[3]).strip() if row[3] else ''
            
            if ne_type:
                mapping[ne_type] = {
                    'domain': domain,
                    'details_file': details_file,
                    'sheet_name': sheet_name
                }
        
        _ne_type_mapping_cache = mapping
        logger.info(f"Загружен маппинг NE Type: {len(mapping)} записей из {mapping_file}")
        return mapping
        
    except Exception as e:
        logger.error(f"Ошибка загрузки маппинга NE Type: {e}")
        return {}


def reload_mapping():
    """Принудительная перезагрузка маппинга NE Type"""
    global _ne_type_mapping_cache
    _ne_type_mapping_cache = None
    return load_ne_type_mapping()

def get_description_file_and_sheet(ne_type):
    """
    Получает имя файла описаний и имя листа для NE типа
    Returns:
        dict: {'domain': str, 'details_file': str, 'sheet_name': str} или None
    """
    mapping = load_ne_type_mapping()
    return mapping.get(ne_type, {})


def _col_letter_to_index(letter):
    """Преобразует букву колонки в индекс (A->1, B->2)"""
    return ord(letter.upper()) - 64


def _find_description_file(filename, network_storage_path, local_path):
    """
    Ищет файл описаний сначала в сетевом хранилище, потом локально
    """
    # 1. Пробуем сетевой путь
    if network_storage_path:
        remote_file = os.path.join(network_storage_path, 'license_details', filename)
        if os.path.exists(remote_file):
            logger.info(f"Файл найден в сетевом хранилище: {remote_file}")
            return remote_file
        else:
            logger.debug(f"Файл не найден в сетевом хранилище: {remote_file}")
    
    # 2. Пробуем локальный путь из конфига
    if local_path:
        local_file = os.path.join(local_path, filename)
        if os.path.exists(local_file):
            logger.info(f"Файл найден локально: {local_file}")
            return local_file
        else:
            logger.debug(f"Файл не найден локально: {local_file}")
    
    # 3. Пробуем относительный путь (в папке проекта)
    project_file = os.path.join('license_details', filename)
    if os.path.exists(project_file):
        logger.info(f"Файл найден в папке проекта: {project_file}")
        return project_file
    
    logger.warning(f"Файл описаний не найден: {filename}")
    return None


def get_capacity_mapping_config(operator_name=None):
    """Возвращает конфиг колонок. formula_col и Sh_col определяются динамически по operator_name"""
    config = {
        'col_key': 'A',
        'col_desc': 'B',
        'col_unit': 'C',
        'col_part': 'D',
        'col_type': 'E',
        'col_parent': 'F',
        'col_spart_coeff': 'G',
        'col_is_main': 'H',
    }
    # formula и sharing определяются динамически из заголовков
    return config


def load_capacity_descriptions(domain, network_storage_path, ne_type=None, operator_name=None):
    """
    Загружает описания CapacityKey из Excel файла для указанного домена
    
    Args:
        domain: str - имя домена (например, 'CScore')
        network_storage_path: str - путь к корню сетевого хранилища
        ne_type: str - тип NE (опционально, для маппинга)
        operator_name: str - имя оператора для поиска колонок Formula_{op}
    
    Returns:
        dict: {capacity_key: {description, unit, part_number, type, parent, formula, sharing, ...}}
    """
    global _capacity_descriptions_cache
    
    if not domain:
        logger.warning("Домен не указан, пропускаем загрузку описаний")
        return {}
    
    # Определяем файл и лист по NE типу, если передан
    details_file = None
    sheet_name = None
    
    if ne_type:
        mapping = get_description_file_and_sheet(ne_type)
        if mapping:
            details_file = mapping.get('details_file')
            sheet_name = mapping.get('sheet_name')
            if mapping.get('domain'):
                domain = mapping['domain']
    
    if not details_file:
        details_file = f'{domain}_license_details.xlsx'
    
    cache_key = f"{domain}_{details_file}_{sheet_name or 'active'}_{operator_name or 'default'}"
    
    if cache_key in _capacity_descriptions_cache:
        return _capacity_descriptions_cache[cache_key]
    
    local_path = None
    try:
        local_path = current_app.config.get('local_license_details_path', '')
    except:
        pass
    
    file_path = _find_description_file(details_file, network_storage_path, local_path)
    
    if not file_path:
        logger.warning(f"Файл описаний не найден для домена {domain}: {details_file}")
        return {}
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        if sheet_name and sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            logger.info(f"Используем лист '{sheet_name}' из {file_path}")
        else:
            sheet = wb.active
            if sheet_name:
                logger.debug(f"Лист '{sheet_name}' не найден, используем активный")
        
        # Читаем заголовки для поиска колонок Formula_{op} и Sh_{op}
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        
        config = get_capacity_mapping_config()
        
        col_idx_key = _col_letter_to_index(config['col_key'])
        col_idx_desc = _col_letter_to_index(config['col_desc'])
        col_idx_unit = _col_letter_to_index(config['col_unit'])
        col_idx_part = _col_letter_to_index(config['col_part'])
        col_idx_type = _col_letter_to_index(config['col_type'])
        col_idx_parent = _col_letter_to_index(config['col_parent'])
        col_idx_spart_coeff = _col_letter_to_index(config['col_spart_coeff'])
        col_idx_is_main = _col_letter_to_index(config['col_is_main'])
        
        # Ищем колонки Formula_{op} и Sh_{op} по имени оператора
        formula_col = None
        Sh_col = None
        if operator_name:
            for i, h in enumerate(headers):
                if h == f'formula_{operator_name}':
                    formula_col = i
                elif h == f'Sh_{operator_name}':
                    Sh_col = i
        
        descriptions = {}
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[col_idx_key - 1]:
                continue
            
            capacity_key = str(row[col_idx_key - 1]).strip()
            description = str(row[col_idx_desc - 1]).strip() if col_idx_desc <= len(row) and row[col_idx_desc - 1] else ''
            unit = str(row[col_idx_unit - 1]).strip() if col_idx_unit <= len(row) and row[col_idx_unit - 1] else ''
            part_number = str(row[col_idx_part - 1]).strip() if col_idx_part <= len(row) and row[col_idx_part - 1] else ''
            key_type = str(row[col_idx_type - 1]).strip().lower() if col_idx_type <= len(row) and row[col_idx_type - 1] else 'bpart'
            parent = str(row[col_idx_parent - 1]).strip() if col_idx_parent <= len(row) and row[col_idx_parent - 1] else ''
            spart_coeff = str(row[col_idx_spart_coeff - 1]).strip() if col_idx_spart_coeff <= len(row) and row[col_idx_spart_coeff - 1] else ''
            is_main_for_spart = str(row[col_idx_is_main - 1]).strip().upper() == 'TRUE' if col_idx_is_main <= len(row) and row[col_idx_is_main - 1] else False
            
            # Читаем formula и sharing для оператора
            formula = ''
            sharing = 1
            if formula_col is not None and formula_col < len(row) and row[formula_col]:
                formula = str(row[formula_col]).strip()
            if Sh_col is not None and Sh_col < len(row) and row[Sh_col]:
                try:
                    sharing = int(row[Sh_col])
                except (ValueError, TypeError):
                    sharing = 1
            
            if capacity_key:
                descriptions[capacity_key] = {
                    'description': description,
                    'unit': unit,
                    'part_number': part_number,
                    'type': key_type,
                    'parent': parent if parent else None,
                    'is_main': key_type == 'main',
                    'valid_date': '',
                    'spart_coeff': spart_coeff if spart_coeff else None,
                    'is_main_for_spart': is_main_for_spart,
                    'formula': formula if formula else None,
                    'sharing': sharing
                }
        
        _capacity_descriptions_cache[cache_key] = descriptions
        logger.info(f"Загружено {len(descriptions)} описаний для домена {domain}, оператор: {operator_name or 'default'}")
        
        return descriptions
        
    except Exception as e:
        logger.error(f"Ошибка загрузки файла {file_path}: {e}")
        return {}
       
def load_full_capacity_list(domain, network_storage_path, ne_type=None, operator_name=None):
    descriptions = load_capacity_descriptions(domain, network_storage_path, ne_type, operator_name)
    
    result = []
    sort_order = 0
    for cap_key, info in descriptions.items():
        item_type = info.get('type', 'bpart')
        
        result.append({
            'name': cap_key,
            'type': item_type,
            'parent': info.get('parent'),
            'part_number': info.get('part_number', ''),
            'dimension': info.get('unit', ''),
            'description': info.get('description', ''),
            'is_main': info.get('is_main', False),
            'is_main_for_spart': info.get('is_main_for_spart', False),
            'is_section': item_type == 'section',
            'spart_coeff': info.get('spart_coeff'),
            'formula': info.get('formula'),
            'sharing': info.get('sharing', 1),
            'sort_order': sort_order
        })
        sort_order += 1
    
    return result

def load_license_structure(domain, network_storage_path, ne_type=None):
    """
    Загружает полную структуру лицензии из Excel файла описаний
    Возвращает список SPart с их BPart (как в файле)
    """
    descriptions = load_capacity_descriptions(domain, network_storage_path, ne_type)
    
    # Группируем по SPart
    sparts_dict = {}
    
    for cap_key, info in descriptions.items():
        key_type = info.get('type', 'bpart')
        parent = info.get('parent')
        
        if key_type == 'spart':
            sparts_dict[cap_key] = {
                'name': cap_key,
                'value': 0,
                'valid_date': '',
                'part_number': info.get('part_number', ''),
                'dimension': info.get('unit', ''),
                'description': info.get('description', ''),
                'children': []
            }
        elif key_type == 'bpart' and parent:
            if parent in sparts_dict:
                sparts_dict[parent]['children'].append({
                    'name': cap_key,
                    'value': 0,
                    'valid_date': '',
                    'part_number': info.get('part_number', ''),
                    'dimension': info.get('unit', ''),
                    'description': info.get('description', ''),
                    'is_main': info.get('is_main', False)
                })
    
    # Сортируем SPart по порядку появления в файле
    return list(sparts_dict.values())

def get_capacity_description(capacity_key, domain, network_storage_path, ne_type=None):
    """Получить описание для конкретного CapacityKey (без привязки к оператору)"""
    if not domain or not capacity_key:
        return None
    descriptions = load_capacity_descriptions(domain, network_storage_path, ne_type)
    return descriptions.get(capacity_key)


def get_all_capacity_descriptions(domain, network_storage_path, ne_type=None):
    """
    Получить все описания для домена
    """
    return load_capacity_descriptions(domain, network_storage_path, ne_type)


def clear_cache():
    """Очистить кэш описаний"""
    global _capacity_descriptions_cache, _ne_type_mapping_cache
    _capacity_descriptions_cache = {}
    _ne_type_mapping_cache = None


def reload_mapping():
    """Принудительная перезагрузка маппинга NE Type"""
    global _ne_type_mapping_cache
    _ne_type_mapping_cache = None
    return load_ne_type_mapping()

def load_targets_from_excel(file_path, operator_name=None):
    """
    Загружает цели из Excel файла targets.xlsx
    Лист Targets: Operator, Domain, Type, Unit, TargetKey, Value_City, Sh_City, ...
    """
    if not os.path.exists(file_path):
        logger.warning(f"Файл целей не найден: {file_path}")
        return [], []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        targets = []
        if 'Targets' in wb.sheetnames:
            sheet = wb['Targets']
            headers = [str(cell.value).strip() if cell.value else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            
            # Находим колонки Value_City и Sh_City
            city_cols = {}  # {col_idx: {'city': 'MSK', 'type': 'value'}}
            for i, h in enumerate(headers):
                if h.startswith('Value_'):
                    city = h.replace('Value_', '').strip()
                    city_cols[i] = {'city': city, 'type': 'value'}
                elif h.startswith('Sh_'):
                    city = h.replace('Sh_', '').strip()
                    # Ищем колонку с таким же городом
                    for j, info in city_cols.items():
                        if info['city'] == city:
                            info['Sh_col'] = i
                            break
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                operator = str(row[0]).strip()
                domain = str(row[1]).strip() if len(row) > 1 else ''
                type_val = str(row[2]).strip() if len(row) > 2 else ''
                # Пропускаем Value_MSK (col 3) и Value_NSK (col 4) - они читаются в цикле city_cols
                unit = str(row[5]).strip() if len(row) > 5 else ''        # Col 5 = Unit
                target_key = str(row[6]).strip() if len(row) > 6 else ''  # Col 6 = TargetKey
                # Col 7 = разделитель "_"
                # Sh_MSK и Sh_NSK читаются в city_cols
                
                if not target_key:
                    continue
                
                for col_idx, info in city_cols.items():
                    if info['type'] == 'value' and col_idx < len(row) and row[col_idx]: 
                        try:
                            value = float(row[col_idx])
                            city = info['city']
                            sharing = 1
                            Sh_col = info.get('Sh_col')
                            if Sh_col and Sh_col < len(row) and row[Sh_col]:
                                sharing = int(row[Sh_col])
                            
                            targets.append({
                                'operator': operator,
                                'domain': domain,
                                'type': type_val,
                                'unit': unit,
                                'target_key': target_key,
                                'city': city,
                                'value': value,
                                'sharing': sharing
                            })
                        except (ValueError, TypeError):
                            pass
        
        logger.info(f"Загружено целей: {len(targets)}")
        return targets
        
    except Exception as e:
        logger.error(f"Ошибка загрузки целей: {e}")
        return [], []
      
def compute_license_targets(targets, capacity_list, operator_name=None):
    """
    Вычисляет целевые значения для каждой комбинации город/NE_type/CapacityKey
    
    Поддерживаемые типы формул (в порядке обработки):
    1. Фиксированные числа: "1", "100", "0.5"
    2. Прямые ссылки на TargetKey: "IMS_Core"
    3. Умножение TargetKey: "PS,vEPC_GGSN_Gbps * 1000"
    4. Переменные (type='variable'): "SP1+SP2+SP3"
    5. Проценты от SPart/переменных: "Total_session * 100%"
    6. Ссылки на другие SPart: "Main_Sessions"
    7. Функции: MAX(), MIN(), AVG(), SUM(), roundup(), if()
    
    Args:
        targets: список словарей из load_targets_from_excel
        capacity_list: список словарей из load_full_capacity_list
        operator_name: str - имя оператора
    
    Returns:
        список словарей [{operator, target_key, city, ne_type, capacity_key, target_value}, ...]
    """
    import re
    import math
    from modules.logger import get_logger
    
    logger = get_logger(__name__)
    results = []
    
    # ========== 1. ПОСТРОЕНИЕ СЛОВАРЯ ЦЕЛЕЙ ИЗ TARGETS.XLSX ==========
    targets_map = {}
    for t in targets:
        key = t['target_key']
        if key not in targets_map:
            targets_map[key] = {}
        targets_map[key][t['city']] = {
            'value': t['value'],
            'sharing': t['sharing'],
            'operator': t['operator']
        }
    
    # Собираем список всех городов
    city_set = set()
    for t in targets:
        city_set.add(t['city'])
    
    if not city_set:
        logger.warning("Нет городов в targets.xlsx")
        return []
    
    # ========== 2. ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========
    
    def get_numeric_value(token, city, targets_map, spart_targets):
        """Получает числовое значение токена (TargetKey, SPart или число)"""
        token_clean = token.strip()
        
        # Проверяем в targets_map
        if token_clean in targets_map and city in targets_map[token_clean]:
            return targets_map[token_clean][city]['value'] / targets_map[token_clean][city]['sharing']
        
        # Проверяем в spart_targets (ключами могут быть с пробелами)
        for spart_name, city_values in spart_targets.items():
            if spart_name.lower() == token_clean.lower():
                return city_values.get(city, 0)
        
        # Пробуем как число
        try:
            return float(token_clean)
        except ValueError:
            return None
    
    def evaluate_simple_expression(expr, city, targets_map, spart_targets):
        """Вычисляет простое арифметическое выражение с + - * /"""
        try:
            # Разбиваем на токены (сохраняя операторы)
            tokens = re.split(r'([+\-*/])', expr)
            result = None
            current_op = '+'
            
            for token in tokens:
                token = token.strip()
                if not token:
                    continue
                
                if token in ['+', '-', '*', '/']:
                    current_op = token
                    continue
                
                value = get_numeric_value(token, city, targets_map, spart_targets)
                if value is None:
                    return None
                
                if result is None:
                    result = value
                elif current_op == '+':
                    result += value
                elif current_op == '-':
                    result -= value
                elif current_op == '*':
                    result *= value
                elif current_op == '/':
                    if value != 0:
                        result /= value
                    else:
                        return None
            
            return result
        except Exception:
            return None
    
    def split_arguments(inner_str):
        """Разбивает аргументы функции, учитывая запятые в именах"""
        if ';' in inner_str:
            return [arg.strip() for arg in inner_str.split(';')]
        if ',' in inner_str:
            return [arg.strip() for arg in inner_str.split(',')]
        return [inner_str.strip()]
    
    def evaluate_condition(condition, city, targets_map, spart_targets):
        """Вычисляет условие для IF"""
        condition_lower = condition.lower().strip()
        
        # AND
        if ' and ' in condition_lower:
            parts = condition_lower.split(' and ')
            for part in parts:
                if not evaluate_condition(part, city, targets_map, spart_targets):
                    return False
            return True
        
        # OR
        if ' or ' in condition_lower:
            parts = condition_lower.split(' or ')
            for part in parts:
                if evaluate_condition(part, city, targets_map, spart_targets):
                    return True
            return False
        
        # Сравнения
        operators = ['>=', '<=', '==', '!=', '>', '<']
        for op in operators:
            if op in condition:
                left, right = condition.split(op, 1)
                left_val = get_numeric_value(left.strip(), city, targets_map, spart_targets)
                right_val = get_numeric_value(right.strip(), city, targets_map, spart_targets)
                
                # Если left_val или right_val None, пробуем найти среди переменных
                if left_val is None:
                    # Ищем среди переменных (ключами могут быть с пробелами)
                    for var_name, city_values in spart_targets.items():
                        if var_name.lower() == left.strip().lower():
                            left_val = city_values.get(city, 0)
                            break
                    if left_val is None:
                        left_val = 0
                
                if right_val is None:
                    for var_name, city_values in spart_targets.items():
                        if var_name.lower() == right.strip().lower():
                            right_val = city_values.get(city, 0)
                            break
                    if right_val is None:
                        right_val = 0
                
                if op == '>':
                    return left_val > right_val
                elif op == '<':
                    return left_val < right_val
                elif op == '>=':
                    return left_val >= right_val
                elif op == '<=':
                    return left_val <= right_val
                elif op == '==':
                    return left_val == right_val
                elif op == '!=':
                    return left_val != right_val
        
        return False
    
    def evaluate_formula(formula_str, city, targets_map, spart_targets):
        """Вычисляет формулу с поддержкой функций"""
        formula_lower = formula_str.lower().strip()
        
        # MAX()
        max_match = re.match(r'max\((.+)\)', formula_lower, re.IGNORECASE)
        if max_match:
            inner = max_match.group(1)
            values = split_arguments(inner)
            max_value = None
            for val_expr in values:
                val = get_numeric_value(val_expr.strip(), city, targets_map, spart_targets)
                if val is not None:
                    if max_value is None or val > max_value:
                        max_value = val
            return max_value
        
        # MIN()
        min_match = re.match(r'min\((.+)\)', formula_lower, re.IGNORECASE)
        if min_match:
            inner = min_match.group(1)
            values = split_arguments(inner)
            min_value = None
            for val_expr in values:
                val = get_numeric_value(val_expr.strip(), city, targets_map, spart_targets)
                if val is not None:
                    if min_value is None or val < min_value:
                        min_value = val
            return min_value
        
        # AVG()
        avg_match = re.match(r'avg\((.+)\)', formula_lower, re.IGNORECASE)
        if avg_match:
            inner = avg_match.group(1)
            values = split_arguments(inner)
            sum_values = 0
            count = 0
            for val_expr in values:
                val = get_numeric_value(val_expr.strip(), city, targets_map, spart_targets)
                if val is not None:
                    sum_values += val
                    count += 1
            return sum_values / count if count > 0 else None
        
        # SUM()
        sum_match = re.match(r'sum\((.+)\)', formula_lower, re.IGNORECASE)
        if sum_match:
            inner = sum_match.group(1)
            values = split_arguments(inner)
            total = 0
            for val_expr in values:
                val = get_numeric_value(val_expr.strip(), city, targets_map, spart_targets)
                if val is not None:
                    total += val
            return total
        
        # IF()
        if_match = re.match(r'if\((.+);\s*(.+);\s*(.+)\)', formula_lower, re.IGNORECASE)
        if not if_match:
            if_match = re.match(r'if\((.+),\s*(.+),\s*(.+)\)', formula_lower, re.IGNORECASE)
        
        if if_match:
            condition = if_match.group(1).strip()
            value_true = if_match.group(2).strip()
            value_false = if_match.group(3).strip()
            
            if evaluate_condition(condition, city, targets_map, spart_targets):
                return get_numeric_value(value_true, city, targets_map, spart_targets)
            else:
                return get_numeric_value(value_false, city, targets_map, spart_targets)
        
        # roundup()
        roundup_match = re.match(r'roundup\((.+),\s*(\d+)\)', formula_lower, re.IGNORECASE)
        if roundup_match:
            inner_expr = roundup_match.group(1).strip()
            decimals = int(roundup_match.group(2))
            value = evaluate_simple_expression(inner_expr, city, targets_map, spart_targets)
            if value is not None:
                multiplier = 10 ** decimals
                return math.ceil(value * multiplier) / multiplier
            return None
        
        # round()
        round_match = re.match(r'round\((.+),\s*(\d+)\)', formula_lower, re.IGNORECASE)
        if round_match:
            inner_expr = round_match.group(1).strip()
            decimals = int(round_match.group(2))
            value = evaluate_simple_expression(inner_expr, city, targets_map, spart_targets)
            if value is not None:
                return round(value, decimals)
            return None
        
        # ceil()
        ceil_match = re.match(r'ceil\((.+)\)', formula_lower, re.IGNORECASE)
        if ceil_match:
            inner_expr = ceil_match.group(1).strip()
            value = evaluate_simple_expression(inner_expr, city, targets_map, spart_targets)
            if value is not None:
                return math.ceil(value)
            return None
        
        # floor()
        floor_match = re.match(r'floor\((.+)\)', formula_lower, re.IGNORECASE)
        if floor_match:
            inner_expr = floor_match.group(1).strip()
            value = evaluate_simple_expression(inner_expr, city, targets_map, spart_targets)
            if value is not None:
                return math.floor(value)
            return None
        
        # Нет функций - простая арифметика
        return evaluate_simple_expression(formula_str, city, targets_map, spart_targets)
    
    # ========== 3. ИНИЦИАЛИЗАЦИЯ ==========
    spart_targets = {}  # {spart_name: {city: target_value}}
    
    # ========== 4. ШАГ 1: ПРЯМЫЕ ССЫЛКИ И ФИКСИРОВАННЫЕ ЧИСЛА ==========
    for item in capacity_list:
        if item['type'] != 'spart':
            continue
        
        formula = item.get('formula')
        if not formula:
            continue
        
        formula_str = str(formula).strip()
        spart_name = item['name']
        ne_type = item.get('ne_type', '')
        
        # Проверяем на фиксированное число
        try:
            fixed_value = float(formula_str)
            spart_targets[spart_name] = {}
            for city in city_set:
                target_value = round(fixed_value, 2)
                spart_targets[spart_name][city] = target_value
                results.append({
                    'operator': operator_name or '',
                    'target_key': formula_str,
                    'city': city,
                    'ne_type': ne_type,
                    'capacity_key': spart_name,
                    'target_value': target_value
                })
            logger.debug(f"Фиксированное число: {spart_name} = {fixed_value}")
            continue
        except (ValueError, TypeError):
            pass
        
        # Проверяем на прямую ссылку на TargetKey
        if formula_str in targets_map:
            spart_targets[spart_name] = {}
            for city, data in targets_map[formula_str].items():
                target_value = round(data['value'] / data['sharing'], 2)
                spart_targets[spart_name][city] = target_value
                results.append({
                    'operator': data['operator'],
                    'target_key': formula_str,
                    'city': city,
                    'ne_type': ne_type,
                    'capacity_key': spart_name,
                    'target_value': target_value
                })
            logger.debug(f"Прямая ссылка: {spart_name} = {formula_str}")
            continue
    
    # ========== 5. ШАГ 2: УМНОЖЕНИЕ TARGETKEY * ЧИСЛО ==========
    for item in capacity_list:
        if item['type'] != 'spart':
            continue
        
        formula = item.get('formula')
        if not formula:
            continue
        
        formula_str = str(formula).strip()
        spart_name = item['name']
        ne_type = item.get('ne_type', '')
        
        if spart_name in spart_targets:
            continue
        
        # Очищаем от пробелов и запятых для поиска
        formula_clean = formula_str.replace(' ', '').replace(',', '')
        
        if '*' in formula_clean and '%' not in formula_str:
            try:
                parts = formula_clean.split('*')
                if len(parts) == 2:
                    target_key_candidate = parts[0]
                    multiplier = float(parts[1])
                    
                    # Ищем TargetKey
                    target_key_found = None
                    if target_key_candidate in targets_map:
                        target_key_found = target_key_candidate
                    else:
                        for tk in targets_map.keys():
                            if tk.replace(',', '').replace(' ', '') == target_key_candidate:
                                target_key_found = tk
                                break
                    
                    if target_key_found:
                        spart_targets[spart_name] = {}
                        for city, data in targets_map[target_key_found].items():
                            target_value = round(data['value'] / data['sharing'] * multiplier, 2)
                            spart_targets[spart_name][city] = target_value
                            results.append({
                                'operator': data['operator'],
                                'target_key': formula_str,
                                'city': city,
                                'ne_type': ne_type,
                                'capacity_key': spart_name,
                                'target_value': target_value
                            })
                        logger.debug(f"Умножение: {spart_name} = {target_key_found} * {multiplier}")
                        continue
            except Exception as e:
                logger.error(f"Ошибка умножения {formula_str}: {e}")
    
    # ========== 6. ШАГ 3: ПЕРЕМЕННЫЕ (ВАЖНО! ДО ПРОЦЕНТОВ) ==========
    variables = {}

    for item in capacity_list:
        if item['type'] != 'variable':
            continue
        
        formula = item.get('formula')
        if not formula:
            continue
        
        formula_str = str(formula).strip()
        var_name = item['name']
        
        # Пропускаем если уже есть
        if var_name in spart_targets:
            continue
        
        # Обработка суммы (SP1+SP2+SP3)
        if '+' in formula_str:
            ref_sparts = [s.strip() for s in formula_str.split('+')]
            
            for city in city_set:
                total = 0
                for ref in ref_sparts:
                    # Ищем в spart_targets (уже вычисленные SPart)
                    if ref in spart_targets and city in spart_targets[ref]:
                        total += spart_targets[ref][city]
                    # Ищем в targets_map
                    elif ref in targets_map and city in targets_map[ref]:
                        total += targets_map[ref][city]['value'] / targets_map[ref][city]['sharing']
                    # Ищем в других переменных
                    elif ref in variables and city in variables[ref]:
                        total += variables[ref][city]
                
                # ВСЕГДА сохраняем переменную, даже если total = 0
                if var_name not in variables:
                    variables[var_name] = {}
                variables[var_name][city] = round(total, 2)
        else:
            # Простая ссылка на один SPart/TargetKey
            ref = formula_str
            for city in city_set:
                value = None
                if ref in spart_targets and city in spart_targets[ref]:
                    value = spart_targets[ref][city]
                elif ref in targets_map and city in targets_map[ref]:
                    value = targets_map[ref][city]['value'] / targets_map[ref][city]['sharing']
                elif ref in variables and city in variables[ref]:
                    value = variables[ref][city]
                else:
                    # Если ссылка нигде не найдена, значение = 0
                    value = 0
                
                # ВСЕГДА сохраняем переменную, даже если value = 0 или None
                if var_name not in variables:
                    variables[var_name] = {}
                variables[var_name][city] = round(value, 2) if value is not None else 0
    
    # Добавляем переменные в spart_targets и результаты
    for var_name, city_values in variables.items():
        if var_name not in spart_targets:
            spart_targets[var_name] = {}
        spart_targets[var_name].update(city_values)
        
        for city, target_value in city_values.items():
            results.append({
                'operator': operator_name or '',
                'target_key': f"variable_{var_name}",
                'city': city,
                'ne_type': '',
                'capacity_key': var_name,
                'target_value': target_value
            })
        logger.debug(f"Переменная: {var_name} = {city_values}")
    
    # ========== 7. ШАГ 4: ПРОЦЕНТЫ ОТ SPART/ПЕРЕМЕННЫХ ==========
    for item in capacity_list:
        if item['type'] != 'spart':
            continue
        
        formula = item.get('formula')
        if not formula:
            continue
        
        formula_str = str(formula).strip()
        spart_name = item['name']
        ne_type = item.get('ne_type', '')
        
        if spart_name in spart_targets:
            continue
        
        if '*' in formula_str and '%' in formula_str:
            try:
                parts = formula_str.split('*')
                ref_spart = parts[0].strip()
                pct_str = parts[1].strip().replace('%', '')
                pct = float(pct_str) / 100
                
                spart_targets[spart_name] = {}
                for city in city_set:
                    ref_value = None
                    # Ищем в spart_targets (включая переменные)
                    if ref_spart in spart_targets and city in spart_targets[ref_spart]:
                        ref_value = spart_targets[ref_spart][city]
                    # Ищем в targets_map
                    elif ref_spart in targets_map and city in targets_map[ref_spart]:
                        ref_value = targets_map[ref_spart][city]['value'] / targets_map[ref_spart][city]['sharing']
                    
                    if ref_value is not None:
                        target_value = round(ref_value * pct, 2)
                        spart_targets[spart_name][city] = target_value
                        results.append({
                            'operator': operator_name or '',
                            'target_key': formula_str,
                            'city': city,
                            'ne_type': ne_type,
                            'capacity_key': spart_name,
                            'target_value': target_value
                        })
                if spart_targets[spart_name]:
                    logger.debug(f"Процент: {spart_name} = {ref_spart} * {pct*100}%")
                continue
            except Exception as e:
                logger.error(f"Ошибка процентов {formula_str}: {e}")
    
    # ========== 8. ШАГ 5: ФОРМУЛЫ С ФУНКЦИЯМИ ==========
    for item in capacity_list:
        if item['type'] != 'spart':
            continue
        
        formula = item.get('formula')
        if not formula:
            continue
        
        formula_str = str(formula).strip()
        spart_name = item['name']
        ne_type = item.get('ne_type', '')
        
        if spart_name in spart_targets:
            continue
        
        # Проверяем наличие функций
        has_function = any(func in formula_str.lower() for func in 
                          ['max(', 'min(', 'avg(', 'sum(', 'if(', 'roundup(', 'round(', 'ceil(', 'floor('])
        
        if has_function:
            for city in city_set:
                value = evaluate_formula(formula_str, city, targets_map, spart_targets)
                if value is not None:
                    if spart_name not in spart_targets:
                        spart_targets[spart_name] = {}
                    target_value = round(value, 2)
                    spart_targets[spart_name][city] = target_value
                    results.append({
                        'operator': operator_name or '',
                        'target_key': formula_str,
                        'city': city,
                        'ne_type': ne_type,
                        'capacity_key': spart_name,
                        'target_value': target_value
                    })
            if spart_targets.get(spart_name):
                logger.debug(f"Функция: {spart_name} = {formula_str}")
    
    # ========== 9. ШАГ 6: ССЫЛКИ НА ДРУГИЕ SPART ==========
    for item in capacity_list:
        if item['type'] != 'spart':
            continue
        
        formula = item.get('formula')
        if not formula:
            continue
        
        formula_str = str(formula).strip()
        spart_name = item['name']
        ne_type = item.get('ne_type', '')
        
        if spart_name in spart_targets:
            continue
        
        # Если формула — просто имя другого SPart или переменной
        if formula_str in spart_targets:
            spart_targets[spart_name] = {}
            for city in city_set:
                if city in spart_targets[formula_str]:
                    target_value = spart_targets[formula_str][city]
                    spart_targets[spart_name][city] = target_value
                    results.append({
                        'operator': operator_name or '',
                        'target_key': formula_str,
                        'city': city,
                        'ne_type': ne_type,
                        'capacity_key': spart_name,
                        'target_value': target_value
                    })
            if spart_targets[spart_name]:
                logger.debug(f"Ссылка: {spart_name} = {formula_str}")
    
    # ========== 10. ШАГ 7: ВИРТУАЛЬНЫЕ SPART ДЛЯ НЕДОСТАЮЩИХ TARGETKEY ==========
    for target_key, cities_data in targets_map.items():
        # Проверяем, есть ли уже SPart
        spart_exists = any(
            item['type'] == 'spart' and item['name'] == target_key 
            for item in capacity_list
        )
        
        if not spart_exists and target_key not in spart_targets:
            spart_targets[target_key] = {}
            for city, data in cities_data.items():
                target_value = round(data['value'] / data['sharing'], 2)
                spart_targets[target_key][city] = target_value
                results.append({
                    'operator': data['operator'],
                    'target_key': target_key,
                    'city': city,
                    'ne_type': '',
                    'capacity_key': target_key,
                    'target_value': target_value
                })
            logger.info(f"Виртуальный SPart: {target_key}")
    
    # ========== 11. СОРТИРОВКА РЕЗУЛЬТАТОВ ==========
    results.sort(key=lambda r: (r['operator'], r['target_key'], r['city'], r['ne_type']))
    
    logger.info(f"Вычислено {len(results)} целевых значений")
    return results


#$