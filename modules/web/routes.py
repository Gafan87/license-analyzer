from flask import Blueprint, render_template, current_app, abort, request, redirect, url_for, jsonify, session, flash
import os
import json
import shutil
import re
from datetime import datetime

import tempfile  # добавь, если нет

from modules.logger import get_logger
from modules.database import (
    init_local_db, get_all_licenses, get_license_by_id, get_change_history, 
    save_base_target, get_base_targets, get_connection, get_filter_options,
    get_dynamic_columns, add_dynamic_column, update_dynamic_column, delete_dynamic_column,
    get_dynamic_values_for_license, update_parsed_cache, get_parsed_cache,
    get_license_aggregated_data, get_all_licenses_with_aggregated,
    get_unique_esn_licenses, get_all_licenses_for_esn, save_license,
    add_tag, get_all_tags, add_tag_to_license, remove_tag_from_license, get_license_tags,
    add_comment, get_comments, save_report_template, get_report_templates, delete_report_template,
    backup_database as backup_db, get_dynamic_columns, get_dynamic_values_for_license
)
from modules.esn_mapper import (
    load_esn_mapping_from_excel, save_esn_mapping_to_db, 
    export_esn_mapping_to_excel, get_mapping_by_esn, get_mapping_by_lsn
)
from modules.scanner import scan_local_folder
from modules.sync_manager import sync_all_licenses, download_db_from_remote, upload_db_to_remote
from modules.target_manager import compare_with_targets, get_targets_for_site
from modules.file_renamer import batch_rename_files, rename_file_by_esn
from modules.excel_importer import import_licenses_from_excel
from modules.tester import run_tests
from modules.excel_evaluator import ExcelFormulaEvaluator
from modules.license_service import LicenseService
from modules.capacity_mapper import get_capacity_description, load_capacity_descriptions

from functools import lru_cache
from datetime import datetime, timedelta

_cache = {}
_cache_time = {}

def get_cached_licenses(operator):
    # Кэш на 5 минут
    now = datetime.now()
    if operator in _cache and (now - _cache_time.get(operator, datetime.min)) < timedelta(minutes=5):
        return _cache[operator]
    
    result = get_unique_esn_licenses(operator)
    _cache[operator] = result
    _cache_time[operator] = now
    return result


logger = get_logger(__name__)
web_bp = Blueprint('web', __name__)


# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========

def get_operator_config(operator_name):
    operators = current_app.config.get('OPERATORS', [])
    for op in operators:
        if op.get('name') == operator_name:
            return op
    return None


# ========== ОСНОВНЫЕ МАРШРУТЫ ==========

@web_bp.route('/')
def index():
    operators = current_app.config.get('OPERATORS', [])
    if operators:
        session['current_operator'] = operators[0]['name']
        return redirect(f'/{operators[0]["name"]}/')
    return "Нет настроенных операторов", 500


@web_bp.route('/<operator>/')
def license_list(operator):
    """Список лицензий (уникальные ESN)"""
    op_config = get_operator_config(operator)
    if not op_config:
        abort(404)
    
    visible_columns = session.get(f'visible_columns_{operator}', 
                                   ['ne_type', 'city', 'site', 'year', 'lsn', 'product', 'version', 'esn', 'tags', 'comments_count'])
    
    licenses = get_unique_esn_licenses(operator)
    
    # Добавляем теги и комментарии
    for lic in licenses:
        agg = get_license_aggregated_data(lic['id'])
        lic['tags'] = agg['tags']
        lic['comments_count'] = agg['comments_count']
    
    ne_types = list(set([l.get('ne_type') for l in licenses if l.get('ne_type')]))
    cities = list(set([l.get('city') for l in licenses if l.get('city')]))
    
    return render_template('index.html',
                          operators=current_app.config['OPERATORS'],
                          current_operator=operator,
                          current_operator_title=op_config.get('title', operator),
                          licenses=licenses,
                          filter_options={'ne_types': ne_types, 'cities': cities},
                          visible_columns=visible_columns)

@web_bp.route('/<operator>/save_columns', methods=['POST'])
def save_visible_columns(operator):
    """Сохраняет настройки видимых колонок и их порядок"""
    from flask import session
    data = request.get_json()
    columns = data.get('columns', [])
    order = data.get('order', columns)
    
    session[f'visible_columns_{operator}'] = columns
    session[f'column_order_{operator}'] = order
    return jsonify({'success': True})

# modules/web/routes.py

@web_bp.route('/<operator>/license/<int:license_id>')
def license_detail(operator, license_id):
    op_config = get_operator_config(operator)
    if not op_config:
        abort(404)

    # Получаем данные лицензии
    license_data = get_license_by_id(license_id)
    if not license_data:
        abort(404)

    # ===== ДОБАВЬТЕ ЭТОТ БЛОК =====
    domain = license_data.get('domain')
    if not domain:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('SELECT domain FROM licenses WHERE id = ?', (license_id,))
        row = cur.fetchone()
        domain = row[0] if row else ''
        conn.close()
        license_data['domain'] = domain
    # ================================

    network_storage_path = current_app.config.get('network_storage_path', '')

    from modules.license_service import LicenseService
    license_data = LicenseService.enrich_resources_with_descriptions(
        license_data, domain, network_storage_path
    )

    license_data['tags'] = get_license_tags(license_id)
    license_data['comments'] = get_comments(license_id)
    license_data['dynamic_values'] = get_dynamic_values_for_license(license_id)

    return render_template('license_detail.html',
                          operators=current_app.config['OPERATORS'],
                          current_operator=operator,
                          current_operator_title=op_config.get('title', operator),
                          license=license_data,
                          all_licenses_for_esn=[license_data],
                          esn=license_data.get('esn'),
                          domain=domain)


# ========== СКАНИРОВАНИЕ Ии СИНХРОНИЗАЦИЯ ==========

@web_bp.route('/<operator>/scan', methods=['POST'])
def scan_operator_route(operator):
    """Запуск сканирования для конкретного оператора"""
    op_config = get_operator_config(operator)
    if not op_config:
        return jsonify({'success': False, 'message': 'Оператор не найден'}), 404
    
    local_path = op_config.get('local_scan_path')
    if not local_path:
        return jsonify({'success': False, 'message': 'Не указан путь для сканирования'}), 400
    
    try:
        licenses = scan_local_folder(local_path, operator)
        
        if not licenses:
            return jsonify({'success': True, 'message': 'Новых лицензий не найдено', 'count': 0})
        
        # Используем пакетное сохранение
        from modules.database import save_licenses_batch
        saved_count, error_count = save_licenses_batch(licenses, operator)
        
        return jsonify({
            'success': True,
            'message': f'Найдено {len(licenses)} лицензий, сохранено: {saved_count}, ошибок: {error_count}',
            'count': len(licenses),
            'saved': saved_count,
            'errors': error_count
        })
    except Exception as e:
        logger.error(f"Ошибка сканирования: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@web_bp.route('/<operator>/sync', methods=['POST'])
def sync_operator_route(operator):
    """Синхронизация лицензий на сервер"""
    op_config = get_operator_config(operator)
    if not op_config:
        return jsonify({'success': False, 'message': 'Оператор не найден'}), 404

    licenses = get_all_licenses(operator=operator)

    if not licenses:
        return jsonify({'success': False, 'message': 'Нет лицензий для синхронизации'}), 400

    remote_base = current_app.config.get('network_storage_path')
    if not remote_base:
        return jsonify({'success': False, 'message': 'Не указан путь к сетевому хранилищу'}), 400

    try:
        licenses_to_sync = []
        for lic in licenses:
            licenses_to_sync.append({
                'filename': lic.get('filename', 'unknown.dat'),
                'operator': operator,
                'domain': lic.get('domain', 'Unknown'),
                'ne_type': lic.get('ne_type', 'Unknown'),
                'city': lic.get('city', 'Unknown'),
                'site': lic.get('site', 'Unknown'),
                'year': lic.get('year', 'permanent'),
                'lsn': lic.get('lsn', ''),
                'local_path': lic.get('local_path', ''),
                'file_hash': lic.get('file_hash', ''),
                'version': lic.get('version', ''),
                'valid_date': lic.get('valid_date', '')
            })

        success, fail = sync_all_licenses(licenses_to_sync, remote_base, modified_by=operator)

        return jsonify({
            'success': True,
            'message': f'Синхронизация завершена: успешно {success}, ошибок {fail}'
        })
    except Exception as e:
        logger.error(f"Ошибка синхронизации: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@web_bp.route('/<operator>/download_db', methods=['POST'])
def download_db_route(operator):
    remote_base = current_app.config.get('network_storage_path')
    local_db = current_app.config.get('local_db_path', 'local_licenses.db')
    network_db = os.path.join(remote_base, 'DB', 'master_licenses.db') if remote_base else None
    
    if not remote_base:
        return jsonify({'success': False, 'message': 'Путь к сетевому хранилищу не задан'}), 400
    
    if download_db_from_remote(network_db, local_db):
        return jsonify({'success': True, 'message': 'База данных загружена'})
    else:
        return jsonify({'success': False, 'message': 'Ошибка загрузки БД'}), 500


@web_bp.route('/<operator>/scan_status', methods=['GET'])
def scan_status(operator):
    count = session.get('scan_count', 0)
    return jsonify({'count': count})


# ========== ОТЧЁТЫ ==========

@web_bp.route('/<operator>/reports')
def reports(operator):
    op_config = get_operator_config(operator)
    if not op_config:
        abort(404)
    
    licenses = get_all_licenses(operator=operator)
    ne_types = list(set([l['ne_type'] for l in licenses if l['ne_type']]))
    cities = list(set([l['city'] for l in licenses if l['city']]))
    
    return render_template('reports.html',
                          operators=current_app.config['OPERATORS'],
                          current_operator=operator,
                          current_operator_title=op_config.get('title', operator),
                          ne_types=ne_types,
                          cities=cities)


@web_bp.route('/<operator>/compare_target', methods=['GET'])
def compare_target(operator):
    op_config = get_operator_config(operator)
    if not op_config:
        abort(404)
    
    ne_type = request.args.get('ne_type')
    city = request.args.get('city')
    site = request.args.get('site')
    year = request.args.get('year')
    
    comparison = None
    if ne_type and city and site and year:
        comparison = compare_with_targets(operator, ne_type, city, site, year)
    
    licenses = get_all_licenses(operator=operator)
    sites = list(set([(l['ne_type'], l['city'], l['site']) for l in licenses if l['site']]))
    
    return render_template('reports.html',
                          operators=current_app.config['OPERATORS'],
                          current_operator=operator,
                          current_operator_title=op_config.get('title', operator),
                          comparison=comparison,
                          sites=sites,
                          selected_ne=ne_type,
                          selected_city=city,
                          selected_site=site,
                          selected_year=year)


# ========== БАЗОВЫЕ ЦЕЛИ ==========

@web_bp.route('/<operator>/base_targets', methods=['GET', 'POST'])
def base_targets(operator):
    op_config = get_operator_config(operator)
    if not op_config:
        abort(404)
    
    ne_type = request.args.get('ne_type')
    city = request.args.get('city')
    site = request.args.get('site')
    
    if request.method == 'POST':
        for key, value in request.form.items():
            if key.startswith('target_') and value:
                cap_key = key[7:]
                save_base_target(operator, ne_type, city, site, cap_key, int(value), operator)
        return redirect(url_for('web.base_targets', operator=operator, ne_type=ne_type, city=city, site=site))
    
    targets = {}
    if ne_type and city and site:
        targets = get_targets_for_site(operator, ne_type, city, site)
    
    licenses = get_all_licenses(operator=operator)
    sites = list(set([(l['ne_type'], l['city'], l['site']) for l in licenses if l['site']]))
    
    return render_template('base_targets.html',
                          operators=current_app.config['OPERATORS'],
                          current_operator=operator,
                          current_operator_title=op_config.get('title', operator),
                          targets=targets,
                          sites=sites,
                          selected_ne=ne_type,
                          selected_city=city,
                          selected_site=site)


# ========== ESN МАППИНГ ==========

@web_bp.route('/<operator>/esn_mapping', methods=['GET', 'POST'])
def esn_mapping(operator):
    op_config = get_operator_config(operator)
    if not op_config:
        abort(404)
    
    if request.method == 'POST':
        if 'upload_file' in request.files:
            file = request.files['upload_file']
            if file and file.filename.endswith('.xlsx'):
                filepath = f"temp_mapping_{operator}.xlsx"
                file.save(filepath)
                mappings = load_esn_mapping_from_excel(filepath)
                save_esn_mapping_to_db(mappings, operator)
                os.remove(filepath)
                flash(f'✅ Загружено {len(mappings)} записей маппинга', 'success')
        
        elif 'export' in request.form:
            remote_base = current_app.config.get('network_storage_path')
            if remote_base:
                export_path = os.path.join(remote_base, 'mapping', f'esn_mapping_{operator}.xlsx')
                os.makedirs(os.path.dirname(export_path), exist_ok=True)
                export_esn_mapping_to_excel(export_path)
                flash(f'✅ Маппинг экспортирован в {export_path}', 'success')
    
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT esn, lsn, operator, domain, ne_type, city, site FROM esn_mapping')
    mappings = cursor.fetchall()
    conn.close()
    
    return render_template('esn_mapping.html',
                          operators=current_app.config['OPERATORS'],
                          current_operator=operator,
                          current_operator_title=op_config.get('title', operator),
                          mappings=mappings)


@web_bp.route('/<operator>/apply_mapping', methods=['POST'])
def apply_mapping_to_licenses(operator):
    """Применяет маппинг ко всем существующим лицензиям"""
    try:
        # Передаём operator в функцию
        updated, conflicts, errors = LicenseService.apply_mapping_batch(operator=operator)
        
        message = f'Обновлено: {updated}, конфликтов (удалено): {conflicts}'
        if errors > 0:
            message += f', ошибок: {errors}'
        
        return jsonify({
            'success': True,
            'updated': updated,
            'conflicts': conflicts,
            'errors': errors,
            'message': message
        })
    except Exception as e:
        logger.error(f"Ошибка применения маппинга: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ========== ИСТОРИЯ ==========

@web_bp.route('/<operator>/history')
def history(operator):
    op_config = get_operator_config(operator)
    if not op_config:
        abort(404)
    
    changes = get_change_history(100)
    
    return render_template('history.html',
                          operators=current_app.config['OPERATORS'],
                          current_operator=operator,
                          current_operator_title=op_config.get('title', operator),
                          changes=changes)


# ========== ИСТЕКАЮЩИЕ ЛИЦЕНЗИИ ==========

def get_expiring_licenses(operator, days=30):
    from datetime import datetime, timedelta
    
    conn = get_connection()
    cursor = conn.cursor()
    
    today = datetime.now().date()
    limit_date = today + timedelta(days=days)
    
    cursor.execute('''
        SELECT l.id, l.ne_type, l.city, l.site, l.year, l.lsn, 
               r.capacity_key, r.value, r.valid_date
        FROM resources r
        JOIN licenses l ON r.license_id = l.id
        WHERE l.operator = ? AND r.valid_date != 'PERMANENT' AND r.valid_date != 'UNKNOWN'
        AND r.valid_date <= ?
        ORDER BY r.valid_date
    ''', (operator, limit_date.isoformat()))
    
    rows = cursor.fetchall()
    conn.close()
    
    expiring = []
    for row in rows:
        try:
            valid_date = datetime.strptime(row[8], '%Y-%m-%d').date()
            days_left = (valid_date - today).days
            if days_left >= 0:
                expiring.append({
                    'license_id': row[0], 'ne_type': row[1], 'city': row[2],
                    'site': row[3], 'year': row[4], 'lsn': row[5],
                    'capacity_key': row[6], 'value': row[7],
                    'valid_date': row[8], 'days_left': days_left
                })
        except:
            pass
    
    return expiring


@web_bp.route('/<operator>/expiring')
def expiring_licenses(operator):
    op_config = get_operator_config(operator)
    if not op_config:
        abort(404)
    
    days = request.args.get('days', 30, type=int)
    expiring = get_expiring_licenses(operator, days)
    
    return render_template('expiring.html',
                          operators=current_app.config['OPERATORS'],
                          current_operator=operator,
                          current_operator_title=op_config.get('title', operator),
                          expiring=expiring,
                          days=days)


# ========== ПЕРЕИМЕНОВАНИЕ ФАЙЛОВ ==========

@web_bp.route('/<operator>/rename_files', methods=['GET', 'POST'])
def rename_files_page(operator):
    op_config = get_operator_config(operator)
    if not op_config:
        abort(404)
    
    incoming_folder = op_config.get('incoming_folder', f'./incoming/{operator}')
    target_folder = current_app.config.get('network_storage_path')
    
    if request.method == 'POST':
        success, failed = batch_rename_files(incoming_folder, target_folder, operator)
        session['renamed_files'] = success
        session.modified = True
        
        return render_template('rename_files.html',
                              operators=current_app.config['OPERATORS'],
                              current_operator=operator,
                              current_operator_title=op_config.get('title', operator),
                              incoming_folder=incoming_folder,
                              success=success,
                              failed=failed)
    
    return render_template('rename_files.html',
                          operators=current_app.config['OPERATORS'],
                          current_operator=operator,
                          current_operator_title=op_config.get('title', operator),
                          incoming_folder=incoming_folder,
                          success=None,
                          failed=None)


@web_bp.route('/<operator>/rename_single', methods=['POST'])
def rename_single_file(operator):
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file'})
    
    file = request.files['file']
    with tempfile.NamedTemporaryFile(suffix=os.path.splitext(file.filename)[1], delete=False) as tmp:
        file.save(tmp.name)
        target_folder = current_app.config.get('network_storage_path')
        success, new_path, message = rename_file_by_esn(tmp.name, target_folder, operator)
        os.unlink(tmp.name)
    
    if success:
        return jsonify({'success': True, 'new_name': os.path.basename(new_path), 'new_path': new_path})
    return jsonify({'success': False, 'error': message})


# ========== ТЕГИ КОММЕНТАРИИ ==========

@web_bp.route('/<operator>/tags/create', methods=['POST'])
def create_tag(operator):
    data = request.get_json()
    name = data.get('name')
    if name:
        tag_id = add_tag(name)
        return jsonify({'success': True, 'tag_id': tag_id})
    return jsonify({'success': False, 'error': 'Имя тега обязательно'})


@web_bp.route('/<operator>/get_all_tags', methods=['GET'])
def get_all_tags_route(operator):
    tags = get_all_tags()
    return jsonify(tags)


@web_bp.route('/<operator>/license/<int:license_id>/tags/add', methods=['POST'])
def add_tag_to_license_route(operator, license_id):
    data = request.get_json()
    tag_id = data.get('tag_id')
    if tag_id:
        add_tag_to_license(license_id, tag_id)
        return jsonify({'success': True})
    return jsonify({'success': False})


@web_bp.route('/<operator>/license/<int:license_id>/tags/remove', methods=['POST'])
def remove_tag_from_license_route(operator, license_id):
    data = request.get_json()
    tag_id = data.get('tag_id')
    if tag_id:
        remove_tag_from_license(license_id, tag_id)
        return jsonify({'success': True})
    return jsonify({'success': False})


@web_bp.route('/<operator>/license/<int:license_id>/comments', methods=['POST'])
def add_comment_route(operator, license_id):
    comment = request.form.get('comment')
    user_name = request.form.get('user_name', operator)
    if comment:
        add_comment(license_id, user_name, comment)
    return redirect(url_for('web.license_detail', operator=operator, license_id=license_id))


# ========== ШАБЛОНЫ ОТЧЁТОВ ==========

@web_bp.route('/<operator>/report_templates', methods=['GET', 'POST'])
def report_templates(operator):
    op_config = get_operator_config(operator)
    if not op_config:
        abort(404)
    
    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description', '')
        filters = {
            'ne_type': request.form.get('filters_ne_type') or '',
            'city': request.form.get('filters_city') or '',
            'status': request.form.get('filters_status') or ''
        }
        columns = request.form.getlist('columns')
        if not columns:
            columns = ['lsn', 'product', 'ne_type', 'city', 'year']
        
        save_report_template(name, description, filters, columns, operator)
        return redirect(url_for('web.report_templates', operator=operator))
    
    templates = get_report_templates()
    licenses = get_all_licenses(operator=operator)
    ne_types = list(set([l['ne_type'] for l in licenses if l['ne_type']]))
    cities = list(set([l['city'] for l in licenses if l['city']]))
    
    return render_template('report_templates.html',
                          operators=current_app.config['OPERATORS'],
                          current_operator=operator,
                          current_operator_title=op_config.get('title', operator),
                          templates=templates,
                          ne_types=ne_types,
                          cities=cities)


@web_bp.route('/<operator>/report_templates/<int:template_id>', methods=['DELETE'])
def delete_report_template_route(operator, template_id):
    delete_report_template(template_id)
    return jsonify({'success': True})


# ========== ИМПОРТ EXCEL ==========

@web_bp.route('/<operator>/import_excel', methods=['POST'])
def import_excel_route(operator):
    import tempfile
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Файл не выбран'})
    
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Требуется Excel файл'})
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        file.save(tmp.name)
        imported, errors = import_licenses_from_excel(tmp.name, operator)
        os.unlink(tmp.name)
    
    return jsonify({
        'success': True,
        'imported': len(imported),
        'errors': errors
    })


# ========== ДАШБОРД ==========

@web_bp.route('/dashboard')
def dashboard():
    operators = current_app.config.get('OPERATORS', [])
    
    stats = []
    total_licenses = 0
    total_expiring = 0
    
    for op in operators:
        licenses = get_all_licenses(operator=op['name'])
        expiring = get_expiring_licenses(op['name'], 30)
        
        stats.append({
            'name': op['title'],
            'id': op['name'],
            'licenses_count': len(licenses),
            'expiring_count': len(expiring),
            'ne_types': len(set([l['ne_type'] for l in licenses if l['ne_type']])),
            'cities': len(set([l['city'] for l in licenses if l['city']]))
        })
        total_licenses += len(licenses)
        total_expiring += len(expiring)
    
    return render_template('dashboard.html',
                          operators=operators,
                          stats=stats,
                          total_licenses=total_licenses,
                          total_expiring=total_expiring)


# ========== НАСТРОЙКИ ==========

@web_bp.route('/settings')
def settings():
    operators = current_app.config.get('OPERATORS', [])
    network_storage = current_app.config.get('network_storage_path', '')
    current_log_level = current_app.config.get('LOG_LEVEL', 'INFO')
    backup_enabled = current_app.config.get('backup_enabled', True)
    backup_interval_days = current_app.config.get('backup_interval_days', 7)
    app_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    licenses = get_all_licenses()
    stats = {
        'licenses_count': len(licenses),
        'operators_count': len(operators),
        'db_size': round(os.path.getsize('local_licenses.db') / (1024*1024), 2) if os.path.exists('local_licenses.db') else 0
    }
    
    current_operator = session.get('current_operator', operators[0]['name'] if operators else 'mts')
    
    return render_template('settings.html',
                          operators=operators,
                          network_storage=network_storage,
                          current_log_level=current_log_level,
                          backup_enabled=backup_enabled,
                          backup_interval_days=backup_interval_days,
                          app_path=app_path,
                          stats=stats,
                          current_operator=current_operator)


@web_bp.route('/settings/update_paths', methods=['POST'])
def update_paths():
    operators = current_app.config.get('OPERATORS', [])
    
    for op in operators:
        new_path = request.form.get(f'path_{op["name"]}')
        if new_path:
            op['local_scan_path'] = new_path
    
    with open('config.json', 'r', encoding='utf-8') as f:
        config = json.load(f)
    config['operators'] = operators
    with open('config.json', 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2)
    
    current_app.config['OPERATORS'] = operators
    flash('✅ Пути сохранены', 'success')
    return redirect(url_for('web.settings'))


@web_bp.route('/settings/update_incoming_paths', methods=['POST'])
def update_incoming_paths():
    operators = current_app.config.get('OPERATORS', [])
    
    for op in operators:
        new_path = request.form.get(f'incoming_{op["name"]}')
        if new_path:
            op['incoming_folder'] = new_path
    
    with open('config.json', 'r', encoding='utf-8') as f:
        config = json.load(f)
    config['operators'] = operators
    with open('config.json', 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2)
    
    current_app.config['OPERATORS'] = operators
    flash('✅ Пути к входящим папкам сохранены', 'success')
    return redirect(url_for('web.settings'))


@web_bp.route('/settings/check_storage')
def check_storage():
    network_storage = current_app.config.get('network_storage_path', '')
    exists = os.path.exists(network_storage) if network_storage else False
    return jsonify({'exists': exists})


@web_bp.route('/settings/backup_list')
def backup_list():
    backup_dir = 'backups'
    backups = []
    if os.path.exists(backup_dir):
        for f in os.listdir(backup_dir):
            if f.endswith('.db'):
                size = round(os.path.getsize(os.path.join(backup_dir, f)) / (1024*1024), 2)
                backups.append({'name': f, 'size': size})
    return jsonify({'backups': sorted(backups, key=lambda x: x['name'], reverse=True)})


@web_bp.route('/settings/restore_backup', methods=['POST'])
def restore_backup():
    filename = request.args.get('file')
    backup_path = os.path.join('backups', filename)
    local_db = current_app.config.get('local_db_path', 'local_licenses.db')
    
    if os.path.exists(backup_path):
        shutil.copy2(backup_path, local_db)
        return jsonify({'success': True, 'message': f'Бэкап {filename} восстановлен'})
    return jsonify({'success': False, 'message': 'Файл не найден'})


@web_bp.route('/settings/system_info')
def system_info():
    import sys
    return jsonify({'python_version': sys.version.split()[0]})


@web_bp.route('/settings/clear_logs', methods=['POST'])
def clear_logs():
    log_dir = 'logs'
    if os.path.exists(log_dir):
        for f in os.listdir(log_dir):
            if f.endswith('.log'):
                os.remove(os.path.join(log_dir, f))
    return jsonify({'success': True, 'message': 'Логи очищены'})


@web_bp.route('/settings/reset_db', methods=['POST'])
def reset_db():
    """Полный сброс локальной БД с пересозданием всех таблиц"""
    from modules.database import reset_and_recreate_db
    
    try:
        reset_and_recreate_db()
        return jsonify({'success': True, 'message': 'БД полностью пересоздана с новой структурой'})
    except Exception as e:
        logger.error(f"Ошибка сброса БД: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@web_bp.route('/settings/log_level', methods=['POST'])
def set_log_level():
    level = request.form.get('level', 'INFO')
    from modules.logger import setup_logging
    setup_logging(level)
    
    with open('config.json', 'r', encoding='utf-8') as f:
        config = json.load(f)
    config['log_level'] = level
    with open('config.json', 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2)
    
    flash(f'✅ Уровень логирования изменён на {level}', 'success')
    return redirect(url_for('web.settings'))


# ========== ТЕСТИРОВАНИЕ ==========

@web_bp.route('/run_tests', methods=['POST'])
def run_system_tests():
    try:
        config = {
            'operators': current_app.config.get('OPERATORS', []),
            'network_storage_path': current_app.config.get('network_storage_path', '')
        }
        result = run_tests(config)
        return jsonify(result)
    except Exception as e:
        logger.error(f"Ошибка при тестировании: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'total': 0,
            'passed': 0,
            'failed': 1,
            'results': []
        }), 500

# ========== API ДЛЯ ESN МАППИНГА (ТЕСТИРОВАНИЕ) ==========

@web_bp.route('/api/esn_mapping', methods=['GET'])
def api_get_esn_mapping():
    """Получить все записи ESN маппинга"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT esn, lsn, operator, domain, ne_type, city, site FROM esn_mapping LIMIT 100')
    rows = cursor.fetchall()
    conn.close()
    
    mappings = [{
        'esn': r[0],
        'lsn': r[1],
        'operator': r[2],
        'domain': r[3],
        'ne_type': r[4],
        'city': r[5],
        'site': r[6]
    } for r in rows]
    
    return jsonify(mappings)


@web_bp.route('/api/esn_mapping/count', methods=['GET'])
def api_get_esn_mapping_count():
    """Получить количество записей в маппинге"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM esn_mapping')
    count = cursor.fetchone()[0]
    conn.close()
    return jsonify({'count': count})


@web_bp.route('/api/apply_mapping', methods=['POST'])
def api_apply_mapping():
    """Применить маппинг и вернуть JSON"""
    op = request.args.get('operator', 'mts')
    try:
        updated, conflicts = LicenseService.apply_mapping_batch(op)
        return jsonify({
            'success': True,
            'updated': updated,
            'conflicts': conflicts,
            'message': f'Обновлено: {updated}, конфликтов: {conflicts}'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@web_bp.route('/api/license_by_esn', methods=['GET'])
def api_get_license_by_esn():
    """Получить лицензию по ESN"""
    esn = request.args.get('esn')
    if not esn:
        return jsonify({'error': 'esn required'}), 400
    
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, lsn, operator, ne_type, city, site, year, product, version
        FROM licenses WHERE esn LIKE ? LIMIT 5
    ''', (f'%{esn}%',))
    rows = cursor.fetchall()
    conn.close()
    
    licenses = [{
        'id': r[0],
        'lsn': r[1],
        'operator': r[2],
        'ne_type': r[3],
        'city': r[4],
        'site': r[5],
        'year': r[6],
        'product': r[7],
        'version': r[8]
    } for r in rows]
    
    return jsonify(licenses)


# ========== ЭКСПОРТ В EXCEL ==========

@web_bp.route('/api/export_licenses', methods=['POST'])
def api_export_licenses():
    data = request.get_json()
    license_ids = data.get('license_ids', [])
    operator = data.get('operator')
    visible_columns = data.get('visible_columns', [])
    
    if license_ids:
        placeholders = ','.join('?' * len(license_ids))
        query = f'''
            SELECT l.id, l.lsn, l.product, l.version, l.ne_type, l.city, l.site, l.year,
                   l.valid_date, l.esn, l.node, l.create_time, l.domain,
                   GROUP_CONCAT(DISTINCT t.name, ', ') as tags,
                   (SELECT COUNT(*) FROM comments c WHERE c.license_id = l.id) as comments_count
            FROM licenses l
            LEFT JOIN license_tags lt ON lt.license_id = l.id
            LEFT JOIN tags t ON lt.tag_id = t.id
            WHERE l.id IN ({placeholders})
            GROUP BY l.id
        '''
        params = license_ids
    else:
        query = '''
            SELECT l.id, l.lsn, l.product, l.version, l.ne_type, l.city, l.site, l.year,
                   l.valid_date, l.esn, l.node, l.create_time, l.domain,
                   GROUP_CONCAT(DISTINCT t.name, ', ') as tags,
                   (SELECT COUNT(*) FROM comments c WHERE c.license_id = l.id) as comments_count
            FROM licenses l
            LEFT JOIN license_tags lt ON lt.license_id = l.id
            LEFT JOIN tags t ON lt.tag_id = t.id
        '''
        params = []
        if operator:
            query += ' WHERE l.operator = ?'
            params.append(operator)
        query += ' GROUP BY l.id ORDER BY l.ne_type, l.city'
    
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    columns_map = {
        'lsn': 'LSN', 'product': 'Продукт', 'version': 'Версия',
        'ne_type': 'NE тип', 'city': 'Город', 'site': 'Сайт',
        'domain': 'Домен', 'year': 'Год', 'valid_date': 'Действует до',
        'esn': 'ESN', 'node': 'Узел', 'create_time': 'Создана',
        'tags': 'Теги', 'comments_count': 'Комментарии'
    }
    
    if not visible_columns:
        visible_columns = list(columns_map.keys())
    
    headers = [columns_map.get(col, col) for col in visible_columns if col in columns_map]
    data_rows = []
    
    for row in rows:
        row_dict = {
            'id': row[0], 'lsn': row[1], 'product': row[2], 'version': row[3],
            'ne_type': row[4], 'city': row[5], 'site': row[6], 'year': row[7],
            'valid_date': row[8], 'esn': row[9], 'node': row[10],
            'create_time': row[11], 'domain': row[12], 'tags': row[13] or '',
            'comments_count': row[14] or 0
        }
        data_rows.append([row_dict.get(col, '') for col in visible_columns if col in columns_map])
    
    return jsonify({
        'headers': headers,
        'rows': data_rows,
        'filename': f'licenses_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    })

# ========== API ДЛЯ ПОЛУЧЕНИЯ РЕСУРСОВ ПО LSN ==========

@web_bp.route('/api/license/<int:license_id>/resources', methods=['GET'])
def api_get_license_resources(license_id):
    from modules.database import get_connection
    from modules.capacity_mapper import get_capacity_description
    from flask import current_app
    import json

    mode = request.args.get('mode', 'total')
    domain = request.args.get('domain', '')
    network_storage_path = current_app.config.get('network_storage_path', '')

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT capacity_key, total_value, permanent_value, dated_values, latest_date, latest_value
        FROM capacity_aggregated WHERE license_id = ?
    ''', (license_id,))
    agg_rows = cursor.fetchall()
    conn.close()

    resources = []

    for row in agg_rows:
        capacity_key = row[0]
        total_value = row[1]
        permanent_value = row[2]
        dated_values_str = row[3]
        latest_date = row[4]
        latest_value = row[5]

        dated_values = []
        if dated_values_str:
            try:
                dated_values = json.loads(dated_values_str)
            except:
                pass

        if mode == 'total':
            value = total_value
            dated_values_list = dated_values if isinstance(dated_values, list) else []
            has_dated = len(dated_values_list) > 0
            has_permanent = permanent_value > 0

            date_parts = []
            if has_permanent:
                date_parts.append('PERMANENT')
            if has_dated:
                latest_date = max([dv['date'] for dv in dated_values_list])
                date_parts.append(latest_date)

            valid_date = ', '.join(date_parts) if date_parts else 'N/A'

        elif mode == 'permanent':
            value = permanent_value if permanent_value > 0 else 0
            valid_date = 'PERMANENT'

        else:  # mode == 'latest'
            latest_dated = None
            latest_dated_value = 0
            for dv in dated_values:
                if latest_dated is None or dv['date'] > latest_dated:
                    latest_dated = dv['date']
                    latest_dated_value = dv['value']
            if latest_dated:
                value = latest_dated_value
                valid_date = latest_dated
            else:
                value = 0
                valid_date = ''

        resources.append({
            'name': capacity_key,
            'value': value,
            'valid_date': valid_date,
            'description': '',
            'unit': ''
        })

    # ========== ОБОГАЩАЕМ ОПИСАНИЯМИ ==========
    for res in resources:
        if res['name']:
            desc = get_capacity_description(res['name'], domain, network_storage_path)
            if desc:
                res['description'] = desc.get('description', '')
                res['unit'] = desc.get('unit', '')
            else:
                res['description'] = ''
                res['unit'] = ''

    return jsonify({'resources': resources})

@web_bp.route('/api/license/<int:license_id>/hierarchy', methods=['GET'])
def api_get_license_hierarchy(license_id):
    from modules.database import get_connection
    from modules.capacity_mapper import load_full_capacity_list
    from flask import current_app

    mode = request.args.get('mode', 'total')
    domain = request.args.get('domain', '')
    show_all = request.args.get('show_all', 'false').lower() == 'true'
    network_storage_path = current_app.config.get('network_storage_path', '')

    # Получаем NE тип лицензии
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT filename FROM licenses WHERE id = ?', (license_id,))
    row = cursor.fetchone()
    filename = row[0] if row else ''
    is_xml = filename.lower().endswith('.xml')
    conn.close()

    if is_xml:
        return get_xml_hierarchy(license_id, mode, domain, show_all_sparts=show_all)
    else:
        return get_dat_hierarchy(license_id, mode, domain)
           
# ========== API ДЛЯ ПРОВЕРКИ ФАЙЛОВ ==========

def get_dat_hierarchy(license_id, mode, domain):
    from modules.database import get_connection
    from modules.capacity_mapper import load_full_capacity_list
    from flask import current_app
    
    network_storage_path = current_app.config.get('network_storage_path', '')
    
    conn = get_connection()
    cursor = conn.cursor()
    
    # Получаем NE тип для определения маппинга
    cursor.execute('SELECT ne_type FROM licenses WHERE id = ?', (license_id,))
    row = cursor.fetchone()
    ne_type = row[0] if row else None
    conn.close()
    
    # Загружаем полный список из маппинга Excel
    full_list = load_full_capacity_list(domain, network_storage_path, ne_type)
    
    # Группируем по SPart
    sparts_dict = {}
    orphans = []
    
    # Создаём SPart
    for item in full_list:
        if item['type'] == 'spart':
            sparts_dict[item['name']] = {
                'name': item['name'],
                'value': 0,
                'valid_date': '',
                'part_number': item.get('part_number', ''),
                'dimension': item.get('dimension', ''),
                'description': item.get('description', ''),
                'children': [],
                'main_bpart': None
            }
    
    # Добавляем BPart к родителям
    for item in full_list:
        if item['type'] == 'bpart':
            parent = item.get('parent')
            if parent and parent in sparts_dict:
                sparts_dict[parent]['children'].append({
                    'name': item['name'],
                    'value': 0,
                    'valid_date': '',
                    'part_number': item.get('part_number', ''),
                    'dimension': item.get('dimension', ''),
                    'description': item.get('description', ''),
                    'is_main': item.get('is_main', False),
                })
                if item.get('is_main_for_spart'):
                    sparts_dict[parent]['main_bpart'] = item['name']
            else:
                orphans.append(item)
        elif item['type'] != 'spart' and not item.get('parent'):
            orphans.append(item)
    
    # Получаем значения из capacity_aggregated
    conn = get_connection()
    cursor = conn.cursor()
    
    if mode == 'total':
        cursor.execute('SELECT capacity_key, total_value, latest_date FROM capacity_aggregated WHERE license_id = ?', (license_id,))
    elif mode == 'permanent':
        cursor.execute('SELECT capacity_key, permanent_value, ? FROM capacity_aggregated WHERE license_id = ? AND permanent_value > 0', ('PERMANENT', license_id))
    else:  # mode == 'latest'
        cursor.execute('SELECT capacity_key, latest_value, latest_date FROM capacity_aggregated WHERE license_id = ? AND latest_date IS NOT NULL', (license_id,))
    
    values_map = {}
    date_map = {}
    for row in cursor.fetchall():
        values_map[row[0]] = row[1] if row[1] else 0
        date_map[row[0]] = row[2] if len(row) > 2 and row[2] else ''
    conn.close()
    
    # Заполняем значениями
    for spart in sparts_dict.values():
        # Сначала ищем прямое значение SPart
        if spart['name'] in values_map:
            spart['value'] = values_map[spart['name']]
            spart['valid_date'] = date_map.get(spart['name'], '')
        else:
            # Ищем значение от main_bpart
            main_bpart = spart.get('main_bpart')
            if main_bpart and main_bpart in values_map:
                spart['value'] = values_map[main_bpart]
                spart['valid_date'] = date_map.get(main_bpart, '')
            else:
                spart['value'] = 0
                spart['valid_date'] = ''
        
        for bpart in spart['children']:
            if bpart['name'] in values_map:
                bpart['value'] = values_map[bpart['name']]
                bpart['valid_date'] = date_map.get(bpart['name'], '')
    
    for orphan in orphans:
        if orphan['name'] in values_map:
            orphan['value'] = values_map[orphan['name']]
            orphan['valid_date'] = date_map.get(orphan['name'], '')
    
    return jsonify({
        'sparts': list(sparts_dict.values()),
        'orphans': orphans
    })

def get_xml_hierarchy(license_id, mode, domain, show_all_sparts=False):
    from modules.database import get_connection
    from modules.capacity_mapper import get_capacity_description, load_full_capacity_list
    from flask import current_app

    network_storage_path = current_app.config.get('network_storage_path', '')

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute('SELECT ne_type FROM licenses WHERE id = ?', (license_id,))
    row = cursor.fetchone()
    ne_type = row[0] if row else None

    full_list = load_full_capacity_list(domain, network_storage_path, ne_type)

    spart_order = {}
    excel_sparts = {}
    sections = []
    
    for item in full_list:
        if item['type'] == 'section':
            sections.append({'name': item['name'], 'position': item['sort_order']})
        elif item['type'] == 'spart':
            order = len(spart_order)
            spart_order[item['name']] = order
            excel_sparts[item['name']] = {
                'name': item['name'],
                'part_number': item.get('part_number', ''),
                'dimension': item.get('dimension', ''),
                'description': item.get('description', ''),
                'main_bpart': None,
                'order': order,
                'section': None
            }
        elif item['type'] == 'bpart' and item.get('is_main_for_spart'):
            parent = item.get('parent')
            if parent and parent in excel_sparts:
                excel_sparts[parent]['main_bpart'] = item['name']

    # Привязка SPart к разделам
    if sections:
        for spart_name, spart_info in excel_sparts.items():
            spart_position = None
            for item in full_list:
                if item['name'] == spart_name:
                    spart_position = item['sort_order']
                    break
            if spart_position is not None:
                prev_section = None
                for section in sections:
                    if section['position'] < spart_position:
                        prev_section = section['name']
                spart_info['section'] = prev_section

    cursor.execute('''
        SELECT id, spart_name, spart_value, spart_valid_date,
               permanent_value, dated_value, dated_date
        FROM license_spart_hierarchy WHERE license_id = ?
        ORDER BY sort_order
    ''', (license_id,))
    sparts = cursor.fetchall()

    existing_sparts = {}
    for spart_id, spart_name, spart_value, spart_valid_date, permanent_value, dated_value, dated_date in sparts:
        existing_sparts[spart_name] = {
            'spart_id': spart_id, 'spart_name': spart_name,
            'spart_value': spart_value, 'spart_valid_date': spart_valid_date,
            'permanent_value': permanent_value, 'dated_value': dated_value,
            'dated_date': dated_date
        }

    if show_all_sparts:
        for spart_name in excel_sparts:
            if spart_name not in existing_sparts:
                existing_sparts[spart_name] = {
                    'spart_id': None, 'spart_name': spart_name,
                    'spart_value': 0, 'spart_valid_date': '',
                    'permanent_value': 0, 'dated_value': 0, 'dated_date': None
                }

    sorted_spart_names = sorted(existing_sparts.keys(), key=lambda n: spart_order.get(n, 9999))
    result_sparts = []
    all_orphan_bparts = []
    current_section = None

    for spart_name in sorted_spart_names:
        spart_data = existing_sparts[spart_name]
        excel_info = excel_sparts.get(spart_name, {})
        
        spart_section = excel_info.get('section')
        if spart_section and spart_section != current_section:
            current_section = spart_section
            result_sparts.append({
                'name': current_section, 'is_section': True,
                'value': '', 'valid_date': '', 'children': [], 'is_empty': False
            })

        spart_id = spart_data['spart_id']
        
        # BPart для этого SPart
        if spart_id is not None:
            cursor.execute('''
                SELECT bpart_name, bpart_value, bpart_valid_date, is_main,
                       permanent_value, dated_value, dated_date
                FROM license_bpart_hierarchy 
                WHERE license_id = ? AND spart_id = ?
                ORDER BY sort_order
            ''', (license_id, spart_id))
            bparts = cursor.fetchall()
        else:
            bparts = []

        # Выбор значения по режиму
        if mode == 'permanent':
            spart_val = spart_data['permanent_value'] if spart_data['permanent_value'] else 0
            spart_date = 'PERMANENT' if spart_data['permanent_value'] and spart_data['permanent_value'] > 0 else ''
        elif mode == 'latest':
            spart_val = spart_data['dated_value'] if spart_data['dated_value'] else 0
            spart_date = spart_data['dated_date'] if spart_data['dated_date'] else ''
        else:
            spart_val = spart_data['spart_value'] if spart_data['spart_value'] else 0
            spart_date = spart_data['spart_valid_date'] if spart_data['spart_valid_date'] else ''

        spart_desc = get_capacity_description(spart_name, domain, network_storage_path)

        bpart_order = {}
        for item in full_list:
            if item['type'] == 'bpart' and item.get('parent') == spart_name:
                bpart_order[item['name']] = len(bpart_order)

        bparts_list = list(bparts)
        bparts_list.sort(key=lambda b: bpart_order.get(b[0], 9999))

        children = []
        for bp_name, bp_value, bp_valid_date, is_main, bp_permanent, bp_dated, bp_dated_date in bparts_list:
            if mode == 'permanent':
                bp_val = bp_permanent if bp_permanent else 0
                bp_date = 'PERMANENT' if bp_permanent and bp_permanent > 0 else ''
            elif mode == 'latest':
                bp_val = bp_dated if bp_dated else 0
                bp_date = bp_dated_date if bp_dated_date else ''
            else:
                bp_val = bp_value if bp_value else 0
                bp_date = bp_valid_date if bp_valid_date else ''

            bp_desc = get_capacity_description(bp_name, domain, network_storage_path)
            children.append({
                'name': bp_name, 'value': bp_val, 'valid_date': bp_date,
                'is_main': bool(is_main),
                'part_number': bp_desc.get('part_number', '') if bp_desc else '',
                'dimension': bp_desc.get('dimension', '') if bp_desc else '',
                'description': bp_desc.get('description', '') if bp_desc else ''
            })

        result_sparts.append({
            'name': spart_name, 'value': spart_val, 'valid_date': spart_date,
            'part_number': spart_desc.get('part_number', '') if spart_desc else excel_info.get('part_number', ''),
            'dimension': spart_desc.get('dimension', '') if spart_desc else excel_info.get('dimension', ''),
            'description': spart_desc.get('description', '') if spart_desc else excel_info.get('description', ''),
            'children': children,
            'main_bpart': excel_info.get('main_bpart'),
            'is_empty': spart_id is None, 'is_section': False
        })

    # Сироты
    cursor.execute('''
        SELECT bpart_name, bpart_value, bpart_valid_date, is_main,
               permanent_value, dated_value, dated_date
        FROM license_bpart_hierarchy 
        WHERE license_id = ? AND spart_id IS NULL
        ORDER BY sort_order
    ''', (license_id,))
    orphans = cursor.fetchall()

    for bp_name, bp_value, bp_valid_date, is_main, bp_permanent, bp_dated, bp_dated_date in orphans:
        if mode == 'permanent':
            bp_val = bp_permanent if bp_permanent else 0
            bp_date = 'PERMANENT' if bp_permanent and bp_permanent > 0 else ''
        elif mode == 'latest':
            bp_val = bp_dated if bp_dated else 0
            bp_date = bp_dated_date if bp_dated_date else ''
        else:
            bp_val = bp_value if bp_value else 0
            bp_date = bp_valid_date if bp_valid_date else ''

        bp_desc = get_capacity_description(bp_name, domain, network_storage_path)
        all_orphan_bparts.append({
            'name': bp_name, 'value': bp_val, 'valid_date': bp_date,
            'is_main': bool(is_main),
            'part_number': bp_desc.get('part_number', '') if bp_desc else '',
            'dimension': bp_desc.get('dimension', '') if bp_desc else '',
            'description': bp_desc.get('description', '') if bp_desc else ''
        })

    conn.close()
    return jsonify({'sparts': result_sparts, 'orphans': all_orphan_bparts})
            
@web_bp.route('/api/check_esn_mapping_files', methods=['GET'])
def api_check_esn_mapping_files():
    """Проверяет наличие файлов ESN маппинга для всех операторов"""
    import os
    from flask import current_app
    
    mapping_path = current_app.config.get('mapping_path', '//A00742028-C5NC5/_Licenses/mapping')
    operators = current_app.config.get('OPERATORS', [])
    
    files = []
    for op in operators:
        mapping_file = op.get('mapping_file', f"{op.get('name')}_esn_mapping.xlsx")
        file_path = os.path.join(mapping_path, mapping_file)
        files.append({
            'name': mapping_file,
            'path': file_path,
            'exists': os.path.exists(file_path)
        })
    
    missing_count = sum(1 for f in files if not f['exists'])
    
    return jsonify({
        'success': True,
        'mapping_path': mapping_path,
        'files': files,
        'missing_count': missing_count,
        'total_count': len(files)
    })


@web_bp.route('/api/check_license_details_files', methods=['GET'])
def api_check_license_details_files():
    """Проверяет наличие файлов описаний CapacityKey для всех NE типов"""
    import os
    from flask import current_app
    
    details_path = current_app.config.get('license_details_path', '//A00742028-C5NC5/_Licenses/license_details')
    
    # Получаем уникальные NE типы из БД
    from modules.database import get_connection
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT DISTINCT ne_type FROM licenses WHERE ne_type IS NOT NULL AND ne_type != ""')
    ne_types = [r[0] for r in cursor.fetchall()]
    conn.close()
    
    files = []
    for ne_type in ne_types:
        file_name = f"{ne_type}_license_details.xlsx"
        file_path = os.path.join(details_path, file_name)
        files.append({
            'name': file_name,
            'path': file_path,
            'exists': os.path.exists(file_path)
        })
    
    # Добавляем общие файлы
    common_files = ['PS_license_details.xlsx', 'vEPC_license_details.xlsx']
    for common_file in common_files:
        file_path = os.path.join(details_path, common_file)
        files.append({
            'name': common_file,
            'path': file_path,
            'exists': os.path.exists(file_path)
        })
    
    missing_count = sum(1 for f in files if not f['exists'])
    
    return jsonify({
        'success': True,
        'details_path': details_path,
        'files': files,
        'missing_count': missing_count,
        'total_count': len(files)
    })
    
@web_bp.route('/api/open_folder')
def api_open_folder():
    """Открыть папку с файлом лицензии (сначала на сервере, потом локально)"""
    import subprocess
    import urllib.parse
    
    license_id = request.args.get('license_id', type=int)
    if not license_id:
        return jsonify({'success': False, 'error': 'Не указан license_id'})
    
    # Получаем данные лицензии
    lic = get_license_by_id(license_id)
    if not lic:
        return jsonify({'success': False, 'error': 'Лицензия не найдена'})
    
    network_storage = current_app.config.get('network_storage_path', '')
    
    # Пытаемся построить сетевой путь
    if network_storage and lic.get('domain') and lic.get('city'):
        # Определяем год как в sync_manager
        valid_date = lic.get('valid_date', '')
        year_folder = lic.get('year', 'permanent')
        if valid_date and valid_date not in ('PERMANENT', 'UNKNOWN'):
            try:
                import re
                date_match = re.match(r'(\d{4})-(\d{2})', valid_date)
                if date_match:
                    year = int(date_match.group(1))
                    month = int(date_match.group(2))
                    year_folder = str(year - 1) if month <= 3 else str(year)
            except:
                pass
        
        # Генерируем имя файла как в sync_manager
        version = lic.get('version', '')
        version_short = version
        if version:
            r_match = re.search(r'R(\d+)', version, re.IGNORECASE)
            if r_match:
                version_short = 'R' + str(int(r_match.group(1)))
        
        filename = f"LIC{lic.get('ne_type', '')}_{version_short}_{lic.get('city', '')}_{lic.get('site', '')}_{year_folder}.dat"
        filename = re.sub(r'[^\w\-.]', '_', filename)
        
        remote_dir = os.path.join(network_storage, lic.get('operator', ''), lic.get('domain', ''), lic.get('city', ''), year_folder)
        remote_path = os.path.join(remote_dir, filename)
        
        # Если сетевая папка существует — открываем её
        if os.path.exists(remote_dir):
            try:
                subprocess.Popen(['explorer', '/select,', os.path.normpath(remote_path)])
                return jsonify({'success': True, 'path': remote_path, 'source': 'network'})
            except Exception as e:
                return jsonify({'success': False, 'error': str(e)})
    
    # Если сетевой путь не найден — пробуем локальный
    local_path = lic.get('local_path')
    if local_path and os.path.exists(local_path):
        try:
            local_dir = os.path.dirname(local_path)
            subprocess.Popen(['explorer', os.path.normpath(local_dir)])
            return jsonify({'success': True, 'path': local_path, 'source': 'local'})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})
    
    return jsonify({'success': False, 'error': 'Файл не найден ни на сервере, ни локально'})
       
@web_bp.route('/api/license/<int:license_id>/info')
def api_get_license_info(license_id):
    """Получить базовую информацию о лицензии"""
    license_data = get_license_by_id(license_id)
    if not license_data:
        return jsonify({'error': 'Not found'}), 404
    
    return jsonify({
        'id': license_data['id'],
        'esn': license_data.get('esn'),
        'lsn': license_data.get('lsn'),
        'product': license_data.get('product'),
        'version': license_data.get('version'),
        'ne_type': license_data.get('ne_type'),
        'city': license_data.get('city'),
        'site': license_data.get('site'),
        'year': license_data.get('year'),
        'valid_date': license_data.get('valid_date'),
        'create_time': license_data.get('create_time'),
        'node': license_data.get('node'),
        'domain': license_data.get('domain'),
        'filename': license_data.get('filename'),
        'operator': license_data.get('operator'),
        'local_path': license_data.get('local_path'),  # ← ДОБАВИТЬ
        'file_hash': license_data.get('file_hash')     # ← ДОБАВИТЬ (для синхронизации)
    })
    
@web_bp.route('/api/license_versions')
def api_get_license_versions():
    """Получить все версии для ESN"""
    operator = request.args.get('operator')
    esn = request.args.get('esn')
    if not operator or not esn:
        return jsonify({'error': 'operator and esn required'}), 400
    
    licenses = get_all_licenses_for_esn(operator, esn)
    
    # Добавляем local_path к каждой версии
    for lic in licenses:
        full = get_license_by_id(lic['id'])
        if full:
            lic['local_path'] = full.get('local_path')
    
    return jsonify(licenses)
  
@web_bp.route('/<operator>/license_by_esn/<esn>')
def license_detail_by_esn(operator, esn):
    """Детальная страница для ESN со всеми LSN"""
    op_config = get_operator_config(operator)
    if not op_config:
        abort(404)
    
    licenses = get_all_licenses_for_esn(operator, esn)
    
    if not licenses:
        abort(404)
    
    # Получаем данные первой лицензии
    main_license = get_license_by_id(licenses[0]['id'])
    
    # Получаем домен и обогащаем ресурсы
    domain = main_license.get('domain')
    if not domain:
        # Если нет в main_license, получи из БД напрямую
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('SELECT domain FROM licenses WHERE id = ?', (licenses[0]['id'],))
        row = cur.fetchone()
        domain = row[0] if row else ''
        conn.close()
    network_storage_path = current_app.config.get('network_storage_path', '')
    
    from modules.license_service import LicenseService
    main_license = LicenseService.enrich_resources_with_descriptions(
        main_license, domain, network_storage_path
    )
    
    # Добавляем теги, комментарии, динамические поля
    main_license['tags'] = get_license_tags(main_license.get('id'))
    main_license['comments'] = get_comments(main_license.get('id'))
    main_license['dynamic_values'] = get_dynamic_values_for_license(main_license['id'])
    
    # Для списка версий тоже обогащаем ресурсы (для предпросмотра)
    all_licenses_data = []
    for lic in licenses:
        full_lic = get_license_by_id(lic['id'])
        if full_lic:
            full_lic = LicenseService.enrich_resources_with_descriptions(
                full_lic, domain, network_storage_path
            )
            all_licenses_data.append(full_lic)
        else:
            all_licenses_data.append(lic)
    print(f"=== ОТЛАДКА ===")
    print(f"operator: {operator}")
    print(f"esn: {esn}")
    print(f"domain: {domain}")
    print(f"network_storage_path: {network_storage_path}")
    print(f"main_license.get('domain'): {main_license.get('domain')}")
    return render_template('license_detail.html',
                          operators=current_app.config['OPERATORS'],
                          current_operator=operator,
                          current_operator_title=op_config.get('title', operator),
                          license=main_license,
                          all_licenses_for_esn=all_licenses_data,
                          esn=esn,
                          network_storage_path=network_storage_path,
                          domain=domain)

@web_bp.route('/<operator>/license_by_esns')
def license_detail_by_esns(operator):
    """Детали лицензии с возможностью переключения между ESN"""
    op_config = get_operator_config(operator)
    if not op_config:
        abort(404)

    esns_param = request.args.get('esns', '')
    esn_list = list(dict.fromkeys([e.strip() for e in esns_param.split(',') if e.strip()]))

    if not esn_list:
        abort(400, description="Не указаны ESN")

    # Получаем первую лицензию для каждого ESN (одна запись на ESN)
    conn = get_connection()
    cursor = conn.cursor()
    placeholders = ','.join('?' * len(esn_list))
    cursor.execute(f'''
        SELECT l.esn, l.ne_type, l.city, l.site, l.domain, l.id
        FROM licenses l
        WHERE l.operator = ? AND l.esn IN ({placeholders})
        AND l.id = (
            SELECT l2.id FROM licenses l2 
            WHERE l2.esn = l.esn AND l2.operator = l.operator
            ORDER BY l2.create_time DESC LIMIT 1
        )
        ORDER BY l.ne_type, l.city, l.esn
    ''', [operator] + esn_list)
    esn_info = cursor.fetchall()
    conn.close()

    # Строим список уникальных ESN
    selected_esns = []
    seen_esns = set()
    for row in esn_info:
        esn = row[0]
        if esn not in seen_esns:
            seen_esns.add(esn)
            selected_esns.append({
                'esn': esn,
                'ne_type': row[1] or '',
                'city': row[2] or '',
                'site': row[3] or '',
                'domain': row[4] or '',
                'license_id': row[5]
            })

    if not selected_esns:
        abort(404, description="Лицензии не найдены")

    # Берём ID из параметра или первую лицензию
    first_license_id = request.args.get('license_id', type=int) or selected_esns[0]['license_id']
    valid_ids = [e['license_id'] for e in selected_esns]
    if first_license_id not in valid_ids:
        first_license_id = selected_esns[0]['license_id']

    # ===== ВАЖНО: Получаем domain из БД как в license_detail_by_esn =====
    main_license = get_license_by_id(first_license_id)
    if not main_license:
        abort(404)

    # Получаем домен — сначала из license_data, потом из БД
    domain = main_license.get('domain')
    if not domain:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('SELECT domain FROM licenses WHERE id = ?', (first_license_id,))
        row = cur.fetchone()
        domain = row[0] if row else ''
        conn.close()

    network_storage_path = current_app.config.get('network_storage_path', '')

    # Обогащаем ресурсы
    from modules.license_service import LicenseService
    main_license = LicenseService.enrich_resources_with_descriptions(
        main_license, domain, network_storage_path
    )

    main_license['tags'] = get_license_tags(first_license_id)
    main_license['comments'] = get_comments(first_license_id)
    main_license['dynamic_values'] = get_dynamic_values_for_license(first_license_id)

    # Получаем все версии для выбранного ESN
    current_esn = main_license.get('esn', esn_list[0])
    all_licenses_for_esn = get_all_licenses_for_esn(operator, current_esn)

    return render_template('license_detail.html',
                          operators=current_app.config['OPERATORS'],
                          current_operator=operator,
                          current_operator_title=op_config.get('title', operator),
                          license=main_license,
                          all_licenses_for_esn=all_licenses_for_esn,
                          esn=current_esn,
                          domain=domain,
                          selected_esns=selected_esns)

# ========== РЕЗЕРВНОЕ КОПИРОВАНИЕ ==========

@web_bp.route('/backup_db', methods=['POST'])
def backup_database_route():
    backup_path = backup_db('all')
    if backup_path:
        return jsonify({'success': True, 'message': f'Бэкап создан: {backup_path}'})
    return jsonify({'success': False, 'message': 'Ошибка создания бэкапа'}), 500


# ========== ЭКСПОРТ В EXCEL ==========

@web_bp.route('/api/export/<operator>', methods=['POST'])
def api_export_to_excel(operator):
    """Экспорт лицензий в Excel"""
    import io
    from modules.database import get_all_licenses_for_export
    
    data = request.get_json()
    license_ids = data.get('license_ids', [])  # Если пусто - экспортируем все
    visible_columns = data.get('visible_columns', [])
    
    # Получаем данные
    licenses = get_all_licenses_for_export(operator, license_ids if license_ids else None)
    
    if not licenses:
        return jsonify({'error': 'Нет данных для экспорта'}), 400
    
    # Формируем заголовки
    headers = []
    col_map = []
    
    column_names = {
        'ne_type': 'NE тип',
        'city': 'Город',
        'site': 'Сайт',
        'year': 'Год',
        'lsn': 'LSN',
        'product': 'Продукт',
        'version': 'Версия',
        'esn': 'ESN',
        'domain': 'Домен',
        'create_time': 'Дата создания',
        'valid_date': 'Действует до',
        'tags_agg': 'Теги',
        'comments_count': 'Комментарии'
    }
    
    for col in visible_columns:
        if col in column_names:
            headers.append(column_names[col])
            col_map.append(col)
        elif col.startswith('dynamic_'):
            # Динамическое поле
            dyn_col = col.replace('dynamic_', '')
            headers.append(dyn_col)
            col_map.append(f'dynamic_{dyn_col}')
    
    # Создаём Excel
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Лицензии {operator}"
    
    # Заголовки
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E60012", end_color="E60012", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Данные
    for row_idx, lic in enumerate(licenses, 2):
        for col_idx, col_name in enumerate(col_map, 1):
            if col_name.startswith('dynamic_'):
                dyn_name = col_name.replace('dynamic_', '')
                value = lic.get('dynamic_values', {}).get(dyn_name, '')
            else:
                value = lic.get(col_name, '')
            
            if col_name == 'comments_count':
                value = f'💬 {value}' if value else ''
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left")
    
    # Авто-ширина
    for col_idx, _ in enumerate(headers, 1):
        max_length = 0
        column = ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)]
        for row_idx in range(1, len(licenses) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        column.width = min(max_length + 2, 50)
    
    # Сохраняем в буфер
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Отправляем файл
    from flask import send_file
    filename = f'licenses_{operator}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ========== ПРЕДУСТАНОВКИ ФИЛЬТРОВ (API для синхронизации с localStorage) ==========

@web_bp.route('/api/presets', methods=['GET'])
def api_get_presets():
    """Получить сохранённые предустановки из localStorage (через запрос)"""
    # Предустановки хранятся в localStorage на клиенте
    # Этот эндпоинт нужен для импорта/экспорта через сервер
    return jsonify({'message': 'Используйте localStorage на клиенте'})


# ========== УМНЫЙ ПОИСК ==========

@web_bp.route('/api/smart_search/<operator>', methods=['POST'])
def api_smart_search(operator):
    """Умный поиск по лицензиям"""
    from modules.database import get_connection
    
    query = request.get_json().get('query', '')
    if not query:
        return jsonify({'results': [], 'suggestions': []})
    
    conn = get_connection()
    cursor = conn.cursor()
    
    suggestions = []
    
    # Поиск CapacityKey
    if any(c.isalpha() for c in query):
        cursor.execute('''
            SELECT DISTINCT capacity_key 
            FROM resources 
            WHERE capacity_key LIKE ? 
            LIMIT 10
        ''', (f'%{query.upper()}%',))
        capacity_keys = [r[0] for r in cursor.fetchall()]
        for key in capacity_keys:
            suggestions.append({
                'text': key,
                'type': 'capacity_key',
                'description': f'CapacityKey: {key}'
            })
    
    # Поиск тегов
    cursor.execute('SELECT name FROM tags WHERE name LIKE ? LIMIT 5', (f'%{query}%',))
    tags = [r[0] for r in cursor.fetchall()]
    for tag in tags:
        suggestions.append({
            'text': f'tag:{tag}',
            'type': 'tag',
            'description': f'Тег: {tag}'
        })
    
    conn.close()
    
    # Команды (подсказки)
    commands = [
        {'text': 'expiring:30', 'type': 'command', 'description': 'Истекает через 30 дней'},
        {'text': 'expiring:60', 'type': 'command', 'description': 'Истекает через 60 дней'},
        {'text': 'permanent', 'type': 'command', 'description': 'Бессрочные лицензии'},
        {'text': 'city:', 'type': 'command', 'description': 'Фильтр по городу (city:Москва)'},
        {'text': 'site:', 'type': 'command', 'description': 'Фильтр по сайту (site:А101)'}
    ]
    
    for cmd in commands:
        if cmd['text'].startswith(query.lower()) or query in cmd['text']:
            suggestions.append(cmd)
    
    return jsonify({'suggestions': suggestions[:10]})


# Добавьте в конец файла modules/web/routes.py

# ========== МАССОВЫЕ ОПЕРАЦИИ С ТЕГАМИ ==========

@web_bp.route('/api/bulk_add_tag', methods=['POST'])
def api_bulk_add_tag():
    """Массовое добавление тега к лицензиям"""
    from modules.database import add_tag, add_tag_to_license, get_connection
    
    data = request.get_json()
    license_ids = data.get('license_ids', [])
    tag_name = data.get('tag_name', '').strip()
    
    if not license_ids:
        return jsonify({'error': 'Нет выбранных лицензий'}), 400
    
    if not tag_name:
        return jsonify({'error': 'Не указан тег'}), 400
    
    # Создаём тег (если не существует)
    tag_id = add_tag(tag_name)
    
    # Добавляем тег к каждой лицензии
    added = 0
    for lic_id in license_ids:
        try:
            add_tag_to_license(int(lic_id), tag_id)
            added += 1
        except Exception as e:
            pass
    
    return jsonify({
        'success': True,
        'added': added,
        'tag_name': tag_name,
        'message': f'Тег "{tag_name}" добавлен к {added} лицензиям'
    })


@web_bp.route('/api/bulk_remove_tag', methods=['POST'])
def api_bulk_remove_tag():
    """Массовое удаление тега с лицензий"""
    from modules.database import remove_tag_from_license, get_connection
    
    data = request.get_json()
    license_ids = data.get('license_ids', [])
    tag_id = data.get('tag_id')
    
    if not license_ids:
        return jsonify({'error': 'Нет выбранных лицензий'}), 400
    
    if not tag_id:
        return jsonify({'error': 'Не указан тег'}), 400
    
    removed = 0
    for lic_id in license_ids:
        try:
            remove_tag_from_license(int(lic_id), int(tag_id))
            removed += 1
        except Exception as e:
            pass
    
    return jsonify({
        'success': True,
        'removed': removed,
        'message': f'Тег удалён с {removed} лицензий'
    })


@web_bp.route('/api/all_tags', methods=['GET'])
def api_all_tags():
    """Получить все теги для массовых операций"""
    from modules.database import get_all_tags
    
    tags = get_all_tags()
    return jsonify(tags)

# ========== ДИНАМИЧЕСКИЕ ПОЛЯ ==========

@web_bp.route('/api/dynamic_columns', methods=['GET'])
def api_list_dynamic_columns():
    """Получить все динамические колонки"""
    from modules.database import get_dynamic_columns
    columns = get_dynamic_columns(active_only=False)
    return jsonify(columns)


@web_bp.route('/api/dynamic_columns', methods=['POST'])
def api_create_dynamic_column():
    """Добавить динамическую колонку"""
    from modules.database import add_dynamic_column, get_dynamic_columns
    
    data = request.get_json()
    column_name = data.get('column_name')
    display_name = data.get('display_name')
    rule_id = data.get('rule_id')
    capacity_key = data.get('capacity_key')
    aggregation_strategy = data.get('aggregation_strategy', 'sum')
    join_separator = data.get('join_separator', ', ')
    
    if not column_name or not display_name:
        return jsonify({'error': 'column_name и display_name обязательны'}), 400
    
    if not rule_id and not capacity_key:
        return jsonify({'error': 'Нужно указать rule_id или capacity_key'}), 400
    
    try:
        column_id = add_dynamic_column(column_name, display_name, rule_id, capacity_key, 
                                        aggregation_strategy, join_separator)
        
        return jsonify({
            'id': column_id,
            'message': f'Колонка добавлена'
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@web_bp.route('/api/dynamic_columns/<int:column_id>', methods=['PUT'])
def api_update_dynamic_column(column_id):
    """Обновить динамическую колонку"""
    from modules.database import update_dynamic_column
    data = request.get_json()
    update_dynamic_column(column_id, **data)
    return jsonify({'message': 'Колонка обновлена'})


@web_bp.route('/api/dynamic_columns/<int:column_id>', methods=['DELETE'])
def api_delete_dynamic_column(column_id):
    """Удалить динамическую колонку"""
    from modules.database import delete_dynamic_column
    delete_dynamic_column(column_id)
    return jsonify({'message': 'Колонка удалена'})


@web_bp.route('/api/extraction_rules', methods=['GET'])
def api_get_extraction_rules():
    """Получить банк правил"""
    rules = current_app.config.get('EXTRACTION_RULES', {'rules': {}})
    return jsonify(rules)


@web_bp.route('/api/extraction_rules', methods=['POST'])
def api_create_extraction_rule():
    """Добавить новое правило в банк"""
    import json
    import re
    import os
    from datetime import datetime
    
    data = request.get_json()
    rule_id = data.get('rule_id')
    display_name = data.get('display_name')
    regex_dat = data.get('regex_dat')
    xpath_xml = data.get('xpath_xml')
    value_type = data.get('value_type', 'text')
    aggregation_strategy = data.get('aggregation_strategy', 'first')
    join_separator = data.get('join_separator', ', ')
    
    if not rule_id or not display_name:
        return jsonify({'error': 'rule_id и display_name обязательны'}), 400
    
    # Валидация regex
    if regex_dat:
        try:
            pattern = regex_dat
            if pattern.startswith('r"') and pattern.endswith('"'):
                pattern = pattern[2:-1]
            elif pattern.startswith("r'") and pattern.endswith("'"):
                pattern = pattern[2:-1]
            re.compile(pattern)
        except re.error as e:
            return jsonify({'error': f'Невалидный regex: {e}'}), 400
    
    # Загружаем текущие правила
    rules_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'extraction_rules.json')
    
    with open(rules_path, 'r', encoding='utf-8') as f:
        rules = json.load(f)
    
    # Добавляем правило
    rules['rules'][rule_id] = {
        'display_name': display_name,
        'regex_dat': regex_dat,
        'xpath_xml': xpath_xml,
        'value_type': value_type,
        'aggregation_strategy': aggregation_strategy,
        'join_separator': join_separator
    }
    rules['updated_at'] = datetime.now().isoformat()
    
    # Сохраняем
    with open(rules_path, 'w', encoding='utf-8') as f:
        json.dump(rules, f, ensure_ascii=False, indent=2)
    
    # Обновляем конфиг приложения
    current_app.config['EXTRACTION_RULES'] = rules
    
    return jsonify({'message': f'Правило {rule_id} добавлено'})


@web_bp.route('/api/capacity_keys/<operator>', methods=['GET'])
def api_get_capacity_keys(operator):
    """Получить список CapacityKey для оператора"""
    from modules.scanner import get_available_capacity_keys
    keys = get_available_capacity_keys(operator)
    return jsonify(keys)


@web_bp.route('/<operator>/dynamic_fields')
def dynamic_fields_page(operator):
    """Страница управления динамическими полями"""
    op_config = get_operator_config(operator)
    if not op_config:
        abort(404)
    
    from modules.database import get_dynamic_columns
    columns = get_dynamic_columns(active_only=False)
    rules = current_app.config.get('EXTRACTION_RULES', {'rules': {}})
    
    return render_template('dynamic_fields.html',
                          operators=current_app.config['OPERATORS'],
                          current_operator=operator,
                          current_operator_title=op_config.get('title', operator),
                          columns=columns,
                          rules=rules.get('rules', {}))
    


