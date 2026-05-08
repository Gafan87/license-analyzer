import openpyxl
import os
import json
from modules.logger import get_logger
from modules.database import get_connection, add_change_history

logger = get_logger(__name__)

def load_esn_mapping_from_excel(file_path):
    """Загружает маппинг ESN из Excel файла"""
    if not os.path.exists(file_path):
        logger.warning(f"Файл маппинга не найден: {file_path}")
        return []
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    mappings = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Пропускаем если ESN пустой
            continue
        
        mapping = {
            'esn': str(row[0]).strip() if row[0] else None,     # Колонка A
            'lsn': str(row[1]).strip() if row[1] else None,     # Колонка B
            'operator': str(row[2]).strip().lower() if row[2] else None,  # Колонка C
            'domain': str(row[3]).strip() if row[3] else None,  # Колонка D (НОВОЕ)
            'ne_type': str(row[4]).strip() if row[4] else None, # Колонка E
            'city': str(row[5]).strip() if row[5] else None,    # Колонка F
            'site': str(row[6]).strip() if row[6] else None     # Колонка G
        }
        mappings.append(mapping)
    
    logger.info(f"Загружено {len(mappings)} записей маппинга из {file_path}")
    return mappings

def save_esn_mapping_to_db(mappings, modified_by='system'):
    """Сохраняет маппинг ESN в БД (7 колонок)"""
    conn = get_connection()
    cursor = conn.cursor()
    
    cursor.execute('DELETE FROM esn_mapping')
    
    added = 0
    for mapping in mappings:
        if mapping.get('esn'):
            try:
                cursor.execute('''
                    INSERT OR REPLACE INTO esn_mapping (esn, lsn, operator, domain, ne_type, city, site)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (mapping['esn'], mapping.get('lsn'), mapping.get('operator'),
                      mapping.get('domain'), mapping.get('ne_type'),
                      mapping.get('city'), mapping.get('site')))
                added += 1
            except Exception as e:
                print(f"Ошибка вставки {mapping.get('esn')}: {e}")
    
    conn.commit()
    conn.close()
    
    add_change_history('esn_mapping', 0, 'REPLACE', 
                      {'count': old_count}, 
                      {'count': added}, 
                      modified_by)
    
    return added

def get_mapping_by_esn(esn):
    """Получает маппинг по ESN (с учётом домена)"""
    if not esn:
        return None
    
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT operator, domain, ne_type, city, site FROM esn_mapping 
        WHERE esn = ? OR esn LIKE ? OR ? LIKE esn
    ''', (esn, f'%{esn}%', esn))
    row = cursor.fetchone()
    conn.close()
    
    if row:
        return {
            'operator': row[0],
            'domain': row[1],
            'ne_type': row[2],
            'city': row[3],
            'site': row[4]
        }
    return None

def get_mapping_by_lsn(lsn):
    """Получает маппинг по LSN (с учётом домена)"""
    if not lsn:
        return None
    
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT operator, domain, ne_type, city, site FROM esn_mapping WHERE lsn = ?', (lsn,))
    row = cursor.fetchone()
    conn.close()
    
    if row:
        return {
            'operator': row[0],
            'domain': row[1],
            'ne_type': row[2],
            'city': row[3],
            'site': row[4]
        }
    return None

def export_esn_mapping_to_excel(file_path):
    """Экспортирует маппинг из БД в Excel (7 колонок)"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT esn, lsn, operator, domain, ne_type, city, site FROM esn_mapping')
    rows = cursor.fetchall()
    conn.close()
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "ESN Mapping"
    
    # Заголовки (7 колонок)
    headers = ['ESN', 'LSN', 'Оператор', 'Домен', 'NE тип', 'Город', 'Сайт']
    sheet.append(headers)
    
    for row in rows:
        # row теперь кортеж из 7 элементов
        sheet.append(list(row))
    
    # Автоширина колонок
    for col_idx, _ in enumerate(headers, 1):
        max_length = 15
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        for row_idx in range(2, len(rows) + 2):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)
    
    wb.save(file_path)
    logger.info(f"Экспортирован маппинг в {file_path}")